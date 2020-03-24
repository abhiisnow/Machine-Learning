# coding: utf-8
from SavewithAllValues import Save_With_All_Values
from New_File_Cleaning import DoCleanInput
import os
import re
import numpy as np
import pandas as pd
from gooey import Gooey, GooeyParser
from nltk.stem import PorterStemmer
from nltk.stem.snowball import SnowballStemmer
from openpyxl import load_workbook
from sklearn.ensemble import RandomForestClassifier
from sklearn.feature_extraction.text import CountVectorizer, TfidfTransformer
from sklearn.linear_model import SGDClassifier
from sklearn.metrics import accuracy_score
from sklearn.metrics import classification_report
from sklearn.metrics import confusion_matrix
from sklearn.model_selection import GridSearchCV
from sklearn.model_selection import train_test_split
from sklearn.naive_bayes import MultinomialNB
from sklearn.pipeline import Pipeline
from sklearn.tree import DecisionTreeClassifier
from termcolor import colored
import camelot
from PyPDF2 import PdfFileReader
import wx
import wx.lib.mixins.listctrl as listmix

st = PorterStemmer()
"""
This class updated all the new values from the prediction file to the training file.
This will work from the second run as there are no available predictions during the first run.
"""
input_frame_path = ''
original_file_path = ''
path_to_save_combined_all_column_output = ''
original_user_frame = pd.DataFrame()
data_from_listctrl = []


class UniqueValuesToTrainingFile:
    def training_the_train_main_method(self, file, column_name, column_name_modified, sheet_name_use):
        """
        :param file: path for the prediction file
        :param column_name: column to update in training file
        :param column_name_modified: column's lable values to be updated
        :param sheet_name_use: sheet name for the predicted file
        :return: message if the file is update or not
        """
        path_to_open_predicted_file = file
        """opening an excel fill and returing a read only file inorder to get the list of columns exisitng"""
        wb = load_workbook(path_to_open_predicted_file, read_only=True)  # open an Excel file and return a workbook
        """checking if column exists in the predicted file
        """
        if column_name in wb.sheetnames:
            path_to_open_training_file = os.getcwd() + '/Data Sets/Training_Files/To_Process_Train_' \
                                         + column_name + '.xlsx'
            """reading the traning file"""
            training_file = pd.read_excel(path_to_open_training_file, sheet_name='Sheet1', encoding='utf-8',
                                          na_filter=False, header=0)
            training_file[column_name] = training_file[column_name].map(lambda x: str(x).strip())
            training_file[column_name] = [word.lower() for word in training_file[column_name]]
            """reading the prediction file"""
            predicted_file = pd.read_excel(path_to_open_predicted_file, sheet_name=sheet_name_use, encoding='utf-8',
                                           na_filter=False,
                                           header=0)
            predicted_file[column_name] = predicted_file[column_name].map(lambda x: x.strip())
            predicted_file[column_name] = [word.lower() for word in predicted_file[column_name]]

            """iterating over the prediction and the training dataframes"""
            for indexes_output_file, rows_predicted_file in predicted_file.iterrows():
                flag = 1
                for indexes_input_file, rows_training_file in training_file.iterrows():
                    """cross checking if there are new values in the prediction file
                    this is done by checking if the each row value exists in both the dataframes or not
                    if yes setting the flag to 0. so that we dont update the new values to training dataframe.
                    else we do update the values"""
                    if rows_predicted_file[column_name] == rows_training_file[column_name]:
                        flag = 0
                        break
                """adding the new values at the end of training dataframe"""
                if flag == 1:
                    training_file = training_file.append({column_name: rows_predicted_file[column_name],
                                                          column_name_modified: rows_predicted_file[
                                                              'Selected_Prediction']}, ignore_index=True)

            """iterating over both the dataframes again 
            but this time check if the values are changed in prediction file"""
            for indexes_input_file, rows_predicted_file in predicted_file.iterrows():
                for indexes_output_file, rows_training_file in training_file.iterrows():

                    """if the column values exists in both the files
                    and if the predicted value is not same as the training dataframe label value
                    then we update the new value to it."""
                    if rows_predicted_file[column_name] == rows_training_file[column_name]:
                        if rows_predicted_file['Selected_Prediction'] != rows_training_file[column_name_modified]:
                            training_file.loc[indexes_output_file, column_name_modified] = rows_predicted_file[
                                'Selected_Prediction']

            path_to_save = os.getcwd() + '/Data Sets/Training_Files/To_Process_Train_' + column_name + '.xlsx'
            """Saving the changes to training file"""
            training_file.to_excel(path_to_save, index=False)
            message = 'Train File Updated'
            print(message)
        else:
            message = 'Train File not updated'

        return message

    def train_the_training_file(self, column_name, input_file_name):
        """
        :param column_name: column to update in training file
        :param input_file_name: input file with path
        :return: message if both file exists or not. if exists then if it is updated or not
        """
        """Getting the folder name where the files are located"""
        input_file_folder_name = input_file_name.rsplit('\\', 1)[1]
        """calling the constructor of traing train file class"""
        train_the_training = UniqueValuesToTrainingFile()
        column_name = column_name.replace(" ", "_")
        column_name_modified = column_name + '_Modified'
        """checking if the training file exists or not
        and also if the prediction file exists or not
        if both exists then we call the method to update the training file"""
        if (os.path.isfile(os.getcwd() + '/Data Sets/Training_Files/To_Process_Train_' + column_name + '.xlsx') and
                os.path.isfile(
                    os.getcwd() + '/Data Sets/Combined_all_column_output/' + input_file_folder_name +
                    '/Predicted_Score_all_columns.xlsx')):
            file = os.getcwd() + '/Data Sets/Combined_all_column_output/' + input_file_folder_name + \
                   '/Predicted_Score_all_columns.xlsx'
            message = train_the_training.training_the_train_main_method(file, column_name, column_name_modified,
                                                                        column_name)
        elif (os.path.isfile(os.getcwd() + '/Data Sets/Training_Files/To_Process_Train_' + column_name + '.xlsx') and
              os.path.isfile(
                  os.getcwd() + '/Data Sets/' + column_name + '_Prediction/Predicted_Score_' + column_name + '.xlsx')):
            file = os.getcwd() + '/Data Sets/' + column_name + '_Prediction/Predicted_Score_' + column_name + '.xlsx'
            message = train_the_training.training_the_train_main_method(file, column_name, column_name_modified,
                                                                        'Sheet1')
        else:
            message = 'Train File or Prediction File Does not exist'
        return print(message)


# all the methods to do text processing are under the below class

"""This calss holds all the methods that are necessary for text processing.
This class is directly called from the main class that checks all the file and then call each method seperately
depending on the need of method."""


class GeneralMethodsClassforTextProcessing:
    # method to create the unique values file and then manually map them to picklist
    def create_unique_values(self, input_data, column_name, column_name_modified, header):
        """
        :param input_data: input file datarame
        :param column_name: column to process
        :param column_name_modified: column with standard values
        :param header: header of the column
        :return:
        """
        """creating empty dataframe with headers"""
        column_name_train = pd.DataFrame(columns=[column_name, column_name_modified])
        """adding the values to column from the input dataframe"""
        column_name_train[column_name] = input_data[header]
        """removing empty values"""
        column_name_train = column_name_train.drop(column_name_train[(column_name_train[column_name] == '')].index)
        """removing all the duplicate values"""
        column_name_train.drop_duplicates(subset=[column_name], keep='first', inplace=True)

        path_to_save = os.getcwd() + '/Data Sets/' + column_name + '_Prediction/Picklist_Train_' + column_name + '.xlsx'
        """saving the unique values into a new file for further processing in next steps"""
        column_name_train.to_excel(path_to_save, index=False)
        return column_name_train.head()

    # to do automatic mapping of train file with pick list

    def map_train_file_with_picklist(self, column_name, column_name_modified, header, picklist_column):
        """
        :param column_name: column to map with picklist
        :param column_name_modified: standard values for each column
        :param header: header of the column
        :param picklist_column: column from the picklist table
        :return:
        """
        """removing unwanted characters from the picklist dataframe headers"""
        picklist_column.columns = [x.strip() for x in picklist_column.columns]
        column_name_copy = column_name + '_Copy'
        """opening a the train file created in last step"""
        path_to_open_train_file = \
            os.getcwd() + '/Data Sets/' + column_name + '_Prediction/Picklist_Train_' + column_name + '.xlsx'
        train_column_name = pd.read_excel(path_to_open_train_file, sheet_name='Sheet1', encoding='utf-8',
                                          na_filter=False, header=0)
        """creating a new empty dataframe with headers"""
        train_column_name_with_picklist = pd.DataFrame(columns=[column_name, column_name_modified, column_name_copy])
        """adding the exisitng train dataframe values to all three columns"""
        train_column_name_with_picklist[column_name] = train_column_name[column_name]
        train_column_name_with_picklist[column_name_modified] = train_column_name[column_name_modified]
        train_column_name_with_picklist[column_name_copy] = train_column_name[column_name]
        """striping spaces"""
        train_column_name_with_picklist[column_name] = train_column_name_with_picklist[column_name].map(
            lambda x: x.strip())
        """dropping empty values"""
        train_column_name_with_picklist = train_column_name_with_picklist.drop(
            train_column_name_with_picklist[(train_column_name_with_picklist[column_name_copy] == '')].index)
        """dropping rows with just space as value"""
        train_column_name_with_picklist = train_column_name_with_picklist.drop(
            train_column_name_with_picklist[(train_column_name_with_picklist[column_name_copy] == ' ')].index)
        """converting to lower case"""
        train_column_name_with_picklist[column_name_copy] = [word.lower() for word in
                                                             train_column_name_with_picklist[column_name_copy]]
        """replacing / with space"""
        train_column_name_with_picklist[column_name_copy] = train_column_name_with_picklist[
            column_name_copy].str.replace('/', ' ')
        """replacing ; with space"""
        train_column_name_with_picklist[column_name_copy] = train_column_name_with_picklist[
            column_name_copy].str.replace(';', ' ')
        """replacing , with space"""
        train_column_name_with_picklist[column_name_copy] = train_column_name_with_picklist[
            column_name_copy].str.replace(',', ' ')
        """replacing alpha numeric and white spaces with single space"""
        train_column_name_with_picklist[column_name_copy] = train_column_name_with_picklist[
            column_name_copy].str.replace('[^\w\s]', ' ')
        """stemming the words"""
        train_column_name_with_picklist[column_name_copy] = train_column_name_with_picklist[column_name_copy].apply(
            lambda x: " ".join([st.stem(word) for word in x.split()]))

        column_name_modified = column_name + '_Modified'
        """creating another dataframe with headers"""
        picklist_column_name = pd.DataFrame(columns=[column_name, column_name_modified])
        picklist_column_name[column_name] = picklist_column[header]
        picklist_column_name[column_name_modified] = picklist_column[header.strip()]

        """processing picklist through the same steps"""
        picklist_column_name = picklist_column_name.drop(
            picklist_column[(picklist_column_name[column_name_modified] == '')].index)
        picklist_column_name[column_name_modified] = picklist_column_name[column_name_modified].map(lambda x: x.strip())
        picklist_column_name[column_name_modified] = [word.lower() for word in
                                                      picklist_column_name[column_name_modified]]
        picklist_column_name[column_name_modified] = picklist_column_name[column_name_modified].str.replace('[^\w\s]',
                                                                                                            ' ')
        picklist_column_name[column_name_modified] = picklist_column_name[column_name_modified].apply(
            lambda x: " ".join([st.stem(word) for word in x.split()]))

        """Here we will do the string similarity
        iterating over the training file and the picklist file after preprocessing steps"""
        for words in train_column_name_with_picklist.itertuples():
            for word in words[3].split():
                for picklist_words in picklist_column_name.itertuples():
                    for picklist_word in picklist_words[2].split():
                        texx = re.search(r"\b{}\b".format(picklist_word), word, re.IGNORECASE) is not None
                        """if the word matches then we add it to the standard values"""
                        if texx:
                            train_column_name_with_picklist.set_value(words.Index, column_name_modified,
                                                                      picklist_words[1])
        """marking the filled and empty values next to each row 
        so that the user can add values to empty values in excel file"""
        for words in train_column_name_with_picklist.itertuples():
            if words[2] == '':
                train_column_name_with_picklist.set_value(words.Index, 'Filled', 'False')
            else:
                train_column_name_with_picklist.set_value(words.Index, 'Filled', 'True')

        """creating a new dataframe with headers and saving the new values in it"""
        save_train_data = pd.DataFrame(columns=[column_name, column_name_modified, 'Filled'])
        save_train_data[column_name] = train_column_name_with_picklist[column_name]
        save_train_data[column_name_modified] = train_column_name_with_picklist[column_name_modified]
        save_train_data['Filled'] = train_column_name_with_picklist['Filled']
        save_train_data.head(20)

        """checking if training directory exists if not then creating a new directory to save the file"""
        if not os.path.exists(os.getcwd() + '/Data Sets/Training_Files'):
            os.makedirs(os.getcwd() + '/Data Sets/Training_Files')
        path_to_save = os.getcwd() + '/Data Sets/Training_Files/To_Process_Train_' + column_name + '.xlsx'
        save_train_data.to_excel(path_to_save, index=False)
        if os.path.isfile(path_to_save):
            print('Training File Exists.')
            pass
        else:
            save_train_data.to_excel(path_to_save, index=False)
        return train_column_name.head()

    # Normalize the count of all the picklist items in the data for more accurate prediction

    def dataset_balancing_in_train_file(self, column_name, column_name_modified):
        """
        :param column_name: Column to process
        :param column_name_modified: Standard values
        :return: First 5 values of the dataframe
        """
        """reading the train file"""
        path_to_open = os.getcwd() + '/Data Sets/Training_Files/To_Process_Train_' + column_name + '.xlsx'
        train_column_name = pd.read_excel(path_to_open, sheet_name='Sheet1', encoding='utf-8', na_filter=False,
                                          header=0)
        """Setting the value in dataframe if it standard value is empty"""
        for index, words in train_column_name.iterrows():
            if words[column_name_modified] == '':
                train_column_name.loc[index, 'Filled'] = 'False'
            else:
                train_column_name.loc[index, 'Filled'] = 'True'

        column_name_used = column_name + '_Used'
        """Grouping by column and the unique values"""
        df = train_column_name.groupby(column_name_modified)[column_name].nunique()
        df_highest = df.nlargest(1)

        df_list = []
        """basic preprocessing"""
        train_column_name[column_name_used] = train_column_name[column_name_modified]
        train_column_name[column_name_used] = train_column_name[column_name_used].str.replace('/', '_')
        train_column_name[column_name_used] = train_column_name[column_name_used].str.replace(';', '_')
        train_column_name[column_name_used] = train_column_name[column_name_used].str.replace(',', '_')
        train_column_name[column_name_used] = train_column_name[column_name_used].str.replace(" ", "")

        count = 0
        """Counting the occurence of each standard values and storing their count in a list"""
        for row in train_column_name.itertuples():
            for i, v in df.iteritems():
                if v <= df_highest.values:
                    if row[2] == i:
                        df_column = 'df_' + str(row[4])
                        if not df_list:
                            count += 1
                            df_list.append(df_column)
                        elif df_column not in df_list:
                            count += 1
                            df_list.append(df_column)
                        else:
                            pass
        """Sorting the training dataframe with standard values"""
        train_column_name_sorted = train_column_name.sort_values(column_name_used)
        df_processed = pd.DataFrame()

        count_df = 0
        """Iterating over the list with the count of each standard value"""
        for items in df_list:
            count_df += 1
            items_df = pd.DataFrame()
            count = 0
            """iterating over the sorted dataframe"""
            for index, row in train_column_name_sorted.iterrows():
                df_val = 'df_' + row[column_name_used]
                """Adding values to different dataframes for each standard values"""
                if df_val == items:
                    items_df.loc[count, column_name] = row[column_name]
                    items_df.loc[count, column_name_modified] = row[column_name_modified]
                    items_df.loc[count, column_name_used] = row[column_name_used]

                    # items_df.set_value(count, column_name, row[1])
                    # items_df.set_value(count, column_name_modified, row[2])
                    # items_df.set_value(count, column_name_used, row[4])
                    count += 1
            """running while all the elements in the dataframe are not read"""
            while len(items_df.index) <= df_highest[0]:
                """Running through different conditions to balance the dataset"""
                if len(items_df.index) <= (df_highest[0] / 2):
                    item_to_append = items_df
                    items_df = pd.concat([items_df, item_to_append], axis=0)
                elif len(items_df.index) > (df_highest[0] / 2):
                    df_to_add = df_highest[0] - len(items_df.index)
                    item_to_append = items_df[-df_to_add:]
                    items_df = pd.concat([items_df, item_to_append], axis=0)
                elif len(items_df.index) == (df_highest[0]):
                    df_to_add = df_highest[0] - len(items_df.index)
                    item_to_append = items_df[-df_to_add:]
                    print(len(item_to_append.index))
                    items_df = pd.concat([items_df, item_to_append], axis=0)
                else:
                    pass
            df_processed = pd.concat([df_processed, items_df], axis=0)
        """Saving the balanced dataset"""
        path_to_save = os.getcwd() + '/Data Sets/' + column_name + '_Prediction/Train_' + column_name + '.xlsx'
        df_processed.to_excel(path_to_save, index=False)
        return df_processed.head()

    # method to map unique values to test dataset

    def create_test_file(self, input_data, column_name, column_name_modified, header):
        """
        :param input_data: input dataframe
        :param column_name: column to process
        :param column_name_modified: standard values
        :param header: Header of the column
        :return: first five values of the dataframe
        """
        """Creating empty test dataframe with headers"""
        test_column_name = pd.DataFrame(columns=[column_name, column_name_modified])
        """Adding values from input dataframe"""
        test_column_name[column_name] = input_data[header]
        """For cases where the column is not in string format, converting to string"""
        test_column_name[column_name] = test_column_name[column_name].astype(str)
        """performing basic preprocessing"""
        test_column_name.rename(columns=lambda x: x.strip())
        test_column_name = test_column_name.drop(test_column_name[(test_column_name[column_name] == '')].index)
        test_column_name = test_column_name.drop(test_column_name[(test_column_name[column_name] == ' ')].index)
        test_column_name.drop_duplicates(subset=[column_name], keep='first', inplace=True)
        test_column_name[column_name] = [word.lower() for word in test_column_name[column_name]]

        """Saving the test dataframe in a excel file"""
        path_to_save = os.getcwd() + '/Data Sets/' + column_name + '_Prediction/Test_' + column_name + '.xlsx'
        test_column_name.to_excel(path_to_save, index=False)
        return test_column_name.head()

    # method to read and clean train and test files.

    def read_train_file(self, column_name):
        """
        :param column_name: column to process
        :return: returning training dataframe
        """
        """reading the train file"""
        path_to_open = os.getcwd() + '/Data Sets/' + column_name + '_Prediction/Train_' + column_name + '.xlsx'
        train_column_name = pd.read_excel(path_to_open, sheet_name='Sheet1', encoding='utf-8', na_filter=False,
                                          header=0)
        """Performing basic preprocessing"""
        train_column_name[column_name] = train_column_name[column_name].map(lambda x: str(x).strip())
        train_column_name[column_name] = [word.lower() for word in train_column_name[column_name]]
        return train_column_name

    def read_test_file(self, column_name):
        """
        :param column_name: column to process
        :return: return test dataframe
        """
        """reading test file into a dataframe"""
        path_to_open = os.getcwd() + '/Data Sets/' + column_name + '_Prediction/Test_' + column_name + '.xlsx'
        test_column_name = pd.read_excel(path_to_open, sheet_name='Sheet1', encoding='utf-8', na_filter=False, header=0)

        """Performing basic preprocessing"""
        test_column_name[column_name] = test_column_name[column_name].apply(str)
        test_column_name.rename(columns=lambda x: x.strip())

        test_column_name[column_name] = test_column_name[column_name].map(lambda x: x.strip())
        test_column_name[column_name] = [word.lower() for word in test_column_name[column_name]]
        test_column_name = test_column_name.drop(test_column_name[(test_column_name[column_name] == '')].index)
        test_column_name = test_column_name.drop(test_column_name[(test_column_name[column_name] == ' ')].index)
        test_column_name.drop_duplicates(subset=[column_name], keep='first', inplace=True)
        return test_column_name

    # Vectorizing the text columns

    def vectorize_data(self, column_name, train_column_name):
        """
        :param column_name: column to process
        :param train_column_name: training dataframe
        :return: returning tfidf transformed column values
        """
        """implementing count vectorizer on training column"""
        count_vect = CountVectorizer()
        x_train_counts = count_vect.fit_transform(train_column_name[column_name])

        """Implementing tfidf transfomer on vectorized column"""
        tfidf_transformer = TfidfTransformer()
        x_train_tfidf = tfidf_transformer.fit_transform(x_train_counts)
        return x_train_tfidf

    # NB model
    def validation_nb_model_fit(self, x_train, y_train, x_val, x_train_tfidf):
        """
        :param x_train: column to train on
        :param y_train: standard values to train on
        :param x_val: validation column
        :param x_train_tfidf: vectorized column
        :return: predictions
        """
        # Pipelining
        """processing the data over the pipeline wih naive bayes"""
        text_clf = Pipeline([('vect', CountVectorizer()), ('tfidf', TfidfTransformer()), ('clf', MultinomialNB()), ])
        text_clf = text_clf.fit(x_train, y_train)
        # using GridSearch CV
        parameters = {'vect__ngram_range': [(1, 1), (1, 2)], 'tfidf__use_idf': (True, False),
                      'clf__alpha': (1e-2, 1e-3), }
        gs_clf = GridSearchCV(text_clf, parameters, n_jobs=-1)
        gs_clf = gs_clf.fit(x_train, y_train)
        predicted = gs_clf.predict(x_val)
        return predicted

    def nb_model_fit(self, column_name, column_name_modified, train_column_name, test_column_name, x_train_tfidf):
        """
        :param column_name: column to process
        :param column_name_modified: standard values
        :param train_column_name: training dataframe
        :param test_column_name: test dataframe
        :param x_train_tfidf: vectorized values
        :return: predictions
        """
        # Pipelining
        """processing the data over the pipeline wih naive bayes"""
        text_clf = Pipeline([('vect', CountVectorizer()), ('tfidf', TfidfTransformer()), ('clf', MultinomialNB()), ])
        text_clf = text_clf.fit(train_column_name[column_name], train_column_name[column_name_modified])
        """Fine tuning parameters"""
        parameters = {'vect__ngram_range': [(1, 1), (1, 2)], 'tfidf__use_idf': (True, False),
                      'clf__alpha': (1e-2, 1e-3), }
        gs_clf = GridSearchCV(text_clf, parameters, n_jobs=-1)
        gs_clf = gs_clf.fit(train_column_name[column_name], train_column_name[column_name_modified])

        predicted = gs_clf.predict(test_column_name[column_name])
        return predicted

    def nb_model_predict(self, column_name, column_name_modified, train_column_name, test_column_name, predicted,
                         all_column_dataframe):
        """
        :param column_name: column to be processed
        :param column_name_modified: standard values
        :param train_column_name: train datframe
        :param test_column_name: test dataframe
        :param predicted: prediction from the model
        :param all_column_dataframe: dataframe with all the columns together
        :return: mean score
        """
        test_column_name['Predict_NB'] = predicted
        save_test = pd.DataFrame()
        all_column_dataframe[column_name] = test_column_name[column_name]
        all_column_dataframe['Predict_NB'] = test_column_name['Predict_NB']
        save_test[column_name] = test_column_name[column_name]
        save_test['Predict_NB'] = test_column_name['Predict_NB']
        path_to_save = os.getcwd() + '/Data Sets/' + column_name + '_Prediction/Predict_NB.xlsx'
        save_test.to_excel(path_to_save, index=False)
        return np.mean(save_test['Predict_NB'] == test_column_name[column_name])

    # Decision Tree model
    def validation_dt_model_fit(self, x_train, y_train, x_val, x_train_tfidf):
        """
        :param x_train:column to train on
        :param y_train:standard values to train on
        :param x_val:validation column
        :param x_train_tfidf:vectorized column
        :return:predictions
        """
        # Pipelining 
        dt = Pipeline([('vect', CountVectorizer()), ('tfidf', TfidfTransformer()),
                       ('clf-dt', DecisionTreeClassifier(criterion="gini", splitter="best",
                                                         max_depth=20, random_state=42)), ])
        _ = dt.fit(x_train, y_train)

        predicted = dt.predict(x_val)
        return predicted

    def dt_model_fit(self, column_name, column_name_modified, train_column_name, test_column_name, x_train_tfidf):
        """
        :param column_name:column to process
        :param column_name_modified:standard values
        :param train_column_name:training dataframe
        :param test_column_name:test dataframe
        :param x_train_tfidf:vectorized values
        :return:predictions
        """
        dt = Pipeline([('vect', CountVectorizer()), ('tfidf', TfidfTransformer()),
                       ('clf-dt', DecisionTreeClassifier(criterion="gini", splitter="best",
                                                         max_depth=30, random_state=50)), ])
        _ = dt.fit(train_column_name[column_name], train_column_name[column_name_modified])
        predicted = dt.predict(test_column_name[column_name])
        return predicted

    def dt_model_predict(self, column_name, column_name_modified, train_column_name, test_column_name, predicted,
                         all_column_dataframe):
        """
        :param column_name:column to be processed
        :param column_name_modified:standard values
        :param train_column_name:train datframe
        :param test_column_name:test dataframe
        :param predicted:prediction from the model
        :param all_column_dataframe:dataframe with all the columns together
        :return:mean score
        """
        test_column_name['Predict_DT'] = predicted
        save_test = pd.DataFrame()
        # Save_Test['Email_Address'] = test_column_name['Email_Address']
        save_test[column_name] = test_column_name[column_name]
        # Save_Test[column_name_modified] = test_column_name[column_name_modified]
        save_test['Predict_DT'] = test_column_name['Predict_DT']
        all_column_dataframe['Predict_DT'] = test_column_name['Predict_DT']
        path_to_save = os.getcwd() + '/Data Sets/' + column_name + '_Prediction/Predict_DT.xlsx'
        save_test.to_excel(path_to_save, index=False)
        return np.mean(save_test['Predict_DT'] == test_column_name[column_name])

    # Random Forest model
    def validation_rf_model_fit(self, x_train, y_train, x_val, x_train_tfidf):
        """
        :param x_train:column to train on
        :param y_train:standard values to train on
        :param x_val:validation column
        :param x_train_tfidf:vectorized column
        :return:predictions
        """
        # Pipelining
        rf = Pipeline([('vect', CountVectorizer()), ('tfidf', TfidfTransformer()),
                       ('clf-rf', RandomForestClassifier(n_estimators=100, max_depth=5, random_state=42)), ])
        _ = rf.fit(x_train, y_train)
        predicted = rf.predict(x_val)
        return predicted

    def rf_model_fit(self, column_name, column_name_modified, train_column_name, test_column_name, X_train_tfidf):
        """
        :param column_name:column to process
        :param column_name_modified:standard values
        :param train_column_name:training dataframe
        :param test_column_name:test dataframe
        :param X_train_tfidf:vectorized values
        :return:predictions
        """
        rf = Pipeline([('vect', CountVectorizer()), ('tfidf', TfidfTransformer()),
                       ('clf-rf', RandomForestClassifier(n_estimators=100, max_depth=5, random_state=42)), ])
        _ = rf.fit(train_column_name[column_name], train_column_name[column_name_modified])
        predicted = rf.predict(test_column_name[column_name])
        return predicted

    def rf_model_predict(self, column_name, column_name_modified, train_column_name, test_column_name, predicted,
                         all_column_dataframe):
        """
        :param column_name:column to be processed
        :param column_name_modified:standard values
        :param train_column_name:train datframe
        :param test_column_name:test dataframe
        :param predicted:prediction from the model
        :param all_column_dataframe:dataframe with all the columns together
        :return:mean score
        """
        test_column_name['Predict_RF'] = predicted
        save_test = pd.DataFrame()
        # Save_Test['Email_Address'] = test_column_name['Email_Address']
        save_test[column_name] = test_column_name[column_name]
        # Save_Test[column_name_modified] = test_column_name[column_name_modified]
        save_test['Predict_RF'] = test_column_name['Predict_RF']
        all_column_dataframe['Predict_RF'] = test_column_name['Predict_RF']
        path_to_save = os.getcwd() + '/Data Sets/' + column_name + '_Prediction/Predict_RF.xlsx'
        save_test.to_excel(path_to_save, index=False)
        return np.mean(save_test['Predict_RF'] == test_column_name[column_name])

    # SVM model
    def validation_svm_model_fit(self, x_train, y_train, x_val, x_train_tfidf):
        """
        :param x_train:column to train on
        :param y_train:standard values to train on
        :param x_val:validation column
        :param x_train_tfidf:vectorized column
        :return:predictions
        """
        # using SVM
        text_clf_svm = Pipeline([('vect', CountVectorizer()), ('tfidf', TfidfTransformer()),
                                 ('clf-svm', SGDClassifier(loss='hinge', penalty='l2', alpha=1e-3, max_iter=10,
                                                           random_state=42)), ])
        _ = text_clf_svm.fit(x_train, y_train)
        predicted_svm = text_clf_svm.predict(x_val)
        return predicted_svm

    def svm_model_fit(self, column_name, column_name_modified, train_column_name, test_column_name, x_train_tfidf):
        """
        :param column_name:column to process
        :param column_name_modified:standard values
        :param train_column_name:training dataframe
        :param test_column_name:test dataframe
        :param x_train_tfidf:vectorized values
        :return:predictions
        """
        # using SVM
        text_clf_svm = Pipeline([('vect', CountVectorizer()), ('tfidf', TfidfTransformer()),
                                 ('clf-svm', SGDClassifier(loss='hinge', penalty='l2', alpha=1e-3, max_iter=10,
                                                           random_state=42)), ])
        _ = text_clf_svm.fit(train_column_name[column_name], train_column_name[column_name_modified])
        predicted_svm = text_clf_svm.predict(test_column_name[column_name])
        return predicted_svm

    def svm_model_predict(self, column_name, column_name_modified, train_column_name, test_column_name, predicted_svm,
                          all_column_dataframe):
        """
        :param column_name:column to be processed
        :param column_name_modified:standard values
        :param train_column_name:train datframe
        :param test_column_name:test dataframe
        :param predicted_svm:prediction from the model
        :param all_column_dataframe:dataframe with all the columns together
        :return:mean score
        """
        test_column_name['Predict_SVM'] = predicted_svm
        save_test = pd.DataFrame()
        # Save_Test['Email_Address'] = test_column_name['Email_Address']
        save_test[column_name] = test_column_name[column_name]
        # Save_Test[column_name_modified] = test_column_name[column_name_modified]
        save_test['Predict_SVM'] = test_column_name['Predict_SVM']
        all_column_dataframe['Predict_SVM'] = test_column_name['Predict_SVM']
        path_to_save = os.getcwd() + '/Data Sets/' + column_name + '_Prediction/Predict_SVM.xlsx'
        save_test.to_excel(path_to_save, index=False)
        return np.mean(save_test['Predict_SVM'] == test_column_name[column_name])

    # SVM GS model
    def validation_svm_grid_search_model_fit(self, x_train, y_train, x_val, x_train_tfidf):
        """
        :param x_train:column to train on
        :param y_train:standard values to train on
        :param x_val:validation column
        :param x_train_tfidf:vectorized column
        :return:predictions
        """
        text_clf_svm = Pipeline([('vect', CountVectorizer()), ('tfidf', TfidfTransformer()),
                                 ('clf-svm', SGDClassifier(loss='hinge', penalty='l2', alpha=1e-3, max_iter=10,
                                                           random_state=42)), ])
        _ = text_clf_svm.fit(x_train, y_train)
        # using Grid Search CV with SVM
        parameters_svm = {'vect__ngram_range': [(1, 1), (1, 2)], 'tfidf__use_idf': (True, False),
                          'clf-svm__alpha': (1e-2, 1e-3), }
        gs_clf_svm = GridSearchCV(text_clf_svm, parameters_svm, n_jobs=-1)
        gs_clf_svm = gs_clf_svm.fit(x_train, y_train)
        predicted_gs_clf_svm = gs_clf_svm.predict(x_val)
        return predicted_gs_clf_svm

    def svm_grid_search_model_fit(self, column_name, column_name_modified,
                                  train_column_name, test_column_name, x_train_tfidf):
        """
        :param column_name:column to process
        :param column_name_modified:standard values
        :param train_column_name:training dataframe
        :param test_column_name:test dataframe
        :param x_train_tfidf:vectorized values
        :return:predictions
        """
        text_clf_svm = Pipeline([('vect', CountVectorizer()), ('tfidf', TfidfTransformer()),
                                 ('clf-svm', SGDClassifier(loss='hinge', penalty='l2', alpha=1e-3, max_iter=10,
                                                           random_state=42)), ])
        _ = text_clf_svm.fit(train_column_name[column_name], train_column_name[column_name_modified])
        # using Grid Search CV with SVM
        parameters_svm = {'vect__ngram_range': [(1, 1), (1, 2)], 'tfidf__use_idf': (True, False),
                          'clf-svm__alpha': (1e-2, 1e-3), }
        gs_clf_svm = GridSearchCV(text_clf_svm, parameters_svm, n_jobs=-1)
        gs_clf_svm = gs_clf_svm.fit(train_column_name[column_name], train_column_name[column_name_modified])
        predicted_gs_clf_svm = gs_clf_svm.predict(test_column_name[column_name])
        return predicted_gs_clf_svm

    def svm_grid_search_model_predict(self, column_name, column_name_modified, train_column_name, test_column_name,
                                      predicted_gs_clf_svm, all_column_dataframe):
        """
        :param column_name:column to be processed
        :param column_name_modified:standard values
        :param train_column_name:train datframe
        :param test_column_name:test dataframe
        :param predicted_gs_clf_svm:prediction from the model
        :param all_column_dataframe:dataframe with all the columns together
        :return:mean score
        """
        test_column_name['Predict_SVM_GS'] = predicted_gs_clf_svm
        save_test = pd.DataFrame()
        # Save_Test['Email_Address'] = test_column_name['Email_Address']
        save_test[column_name] = test_column_name[column_name]
        # Save_Test[column_name_modified] = test_column_name[column_name_modified]
        save_test['Predict_SVM_GS'] = test_column_name['Predict_SVM_GS']
        all_column_dataframe['Predict_SVM_GS'] = test_column_name['Predict_SVM_GS']
        path_to_save = os.getcwd() + '/Data Sets/' + column_name + '_Prediction/Predict_SVM_GS.xlsx'
        save_test.to_excel(path_to_save, index=False)
        return np.mean(save_test['Predict_SVM_GS'] == test_column_name[column_name])

    # nb nltk model
    def validation_nltk_nb_model_fit(self, x_train, y_train, x_val, x_train_tfidf):
        """
        :param x_train:column to train on
        :param y_train:standard values to train on
        :param x_val:validation column
        :param x_train_tfidf:vectorized column
        :return:predictions
        """
        # nltk.download()
        stemmer = SnowballStemmer("english", ignore_stopwords=True)

        class StemmedCountVectorizer(CountVectorizer):
            def build_analyzer(self):
                analyzer = super(StemmedCountVectorizer, self).build_analyzer()
                return lambda doc: ([stemmer.stem(w) for w in analyzer(doc)])

        stemmed_count_vect = StemmedCountVectorizer(stop_words='english')
        text_mnb_stemmed = Pipeline(
            [('vect', stemmed_count_vect), ('tfidf', TfidfTransformer()), ('mnb', MultinomialNB(fit_prior=False)), ])
        text_mnb_stemmed = text_mnb_stemmed.fit(x_train, y_train)
        predicted_mnb_stemmed = text_mnb_stemmed.predict(x_val)
        return predicted_mnb_stemmed

    def nltk_nb_model_fit(self, column_name, column_name_modified, train_column_name, test_column_name, x_train_tfidf):
        """
        :param column_name:column to process
        :param column_name_modified:standard values
        :param train_column_name:training dataframe
        :param test_column_name:test dataframe
        :param x_train_tfidf:vectorized values
        :return:predictions
        """
        # nltk.download()
        stemmer = SnowballStemmer("english", ignore_stopwords=True)

        class StemmedCountVectorizer(CountVectorizer):
            def build_analyzer(self):
                analyzer = super(StemmedCountVectorizer, self).build_analyzer()
                return lambda doc: ([stemmer.stem(w) for w in analyzer(doc)])

        stemmed_count_vect = StemmedCountVectorizer(stop_words='english')
        text_mnb_stemmed = Pipeline(
            [('vect', stemmed_count_vect), ('tfidf', TfidfTransformer()), ('mnb', MultinomialNB(fit_prior=False)), ])
        text_mnb_stemmed = text_mnb_stemmed.fit(train_column_name[column_name], train_column_name[column_name_modified])
        predicted_mnb_stemmed = text_mnb_stemmed.predict(test_column_name[column_name])
        return predicted_mnb_stemmed

    def nltk_nb_model_predict(self, column_name, column_name_modified, train_column_name, test_column_name,
                              predicted_mnb_stemmed, all_column_dataframe):
        """
        :param column_name:column to be processed
        :param column_name_modified:standard values
        :param train_column_name:train datframe
        :param test_column_name:test dataframe
        :param predicted_mnb_stemmed:prediction from the model
        :param all_column_dataframe:dataframe with all the columns together
        :return:mean score
        """
        test_column_name['Predict_NLTK_NB'] = predicted_mnb_stemmed
        save_test = pd.DataFrame()
        # Save_Test['Email_Address'] = test_column_name['Email_Address']
        save_test[column_name] = test_column_name[column_name]
        # Save_Test[column_name_modified] = test_column_name[column_name_modified]
        save_test['Predict_NLTK_NB'] = test_column_name['Predict_NLTK_NB']
        all_column_dataframe['Predict_NLTK_NB'] = test_column_name['Predict_NLTK_NB']
        path_to_save = os.getcwd() + '/Data Sets/' + column_name + '_Prediction/Predict_NLTK_NB.xlsx'
        save_test.to_excel(path_to_save, index=False)

        return np.mean(save_test['Predict_NLTK_NB'] == test_column_name[column_name])

    # generating classification report for the above data
    def classification_report_nb(self, y_val, predicted):  # target_name_for_classification,
        """
        :param y_val: actual value from test dataset
        :param predicted: predicted value from test dataset
        :return: accuracy score
        """
        precision_nb = accuracy_score(y_val, predicted)
        return precision_nb

    def classification_report_svm(self, y_val, predicted_svm):  # target_name_for_classification,
        """
        :param y_val: actual value from test dataset
        :param predicted_svm: predicted value from test dataset
        :return:accuracy score
        """
        precision_svm = accuracy_score(y_val, predicted_svm)
        return precision_svm

    def classification_report_gs_svm(self, y_val, predicted_gs_clf_svm):  # target_name_for_classification,
        """
        :param y_val: actual value from test dataset
        :param predicted_gs_clf_svm:predicted value from test dataset
        :return:accuracy score
        """
        precision_gs_svm = accuracy_score(y_val, predicted_gs_clf_svm)
        return precision_gs_svm

    def classification_report_nltk_nb(self, y_val, predicted_mnb_stemmed):  # target_name_for_classification,
        """
        :param y_val: actual value from test dataset
        :param predicted_mnb_stemmed:predicted value from test dataset
        :return:accuracy score
        """
        precision_nltk_nb = accuracy_score(y_val, predicted_mnb_stemmed)
        return precision_nltk_nb

    def classification_report_rf(self, y_val, predicted_rf):  # target_name_for_classification,
        """
        :param y_val: actual value from test dataset
        :param predicted_rf: predicted value from test dataset
        :return: accuracy score
        """
        precision_rf = accuracy_score(y_val, predicted_rf)
        return precision_rf

    def classification_report_dt(self, y_val, predicted_dt):  # target_name_for_classification,
        """
        :param y_val: actual value from test dataset
        :param predicted_dt: predicted value from test dataset
        :return: accuracy score
        """
        precision_dt = accuracy_score(y_val, predicted_dt)
        return precision_dt

    # creating confusion matrix to check the quality of model

    def confusion_matrix_nb(self, column_name, column_name_modified, train_column_name,
                            test_column_name, target_name_for_classification, predicted):
        """
        :param column_name: column trained on
        :param column_name_modified:standard values to train on
        :param train_column_name:column name trained
        :param test_column_name:column name for test
        :param target_name_for_classification:labels used for classification
        :param predicted:predicted values
        :return:confusion matrix with positives and negatives
        """
        return confusion_matrix(test_column_name[column_name_modified], predicted,
                                labels=target_name_for_classification)

    def confusion_matrix_svm(self, column_name, column_name_modified, train_column_name,
                             test_column_name, target_name_for_classification, predicted_svm):
        """
        :param column_name: column trained on
        :param column_name_modified: standard values to train on
        :param train_column_name: column name trained
        :param test_column_name: column name for test
        :param target_name_for_classification:labels used for classification
        :param predicted_svm: predicted values
        :return:confusion matrix with positives and negatives
        """
        return confusion_matrix(test_column_name[column_name_modified], predicted_svm,
                                labels=target_name_for_classification)

    def confusion_matrix_gs_svm(self, column_name, column_name_modified, train_column_name,
                                test_column_name, target_name_for_classification, predicted_gs_clf_svm):
        """
        :param column_name: column trained on
        :param column_name_modified: standard values to train on
        :param train_column_name: column name trained
        :param test_column_name: column name for test
        :param target_name_for_classification: labels used for classification
        :param predicted_gs_clf_svm: predicted values
        :return: confusion matrix with positives and negatives
        """
        return confusion_matrix(test_column_name[column_name_modified], predicted_gs_clf_svm,
                                labels=target_name_for_classification)

    def confusion_matrix_nltk_nb(self, column_name, column_name_modified, train_column_name,
                                 test_column_name, target_name_for_classification, predicted_mnb_stemmed):
        """
        :param column_name: column trained on
        :param column_name_modified: standard values to train on
        :param train_column_name: column name trained
        :param test_column_name: column name for test
        :param target_name_for_classification: labels used for classification
        :param predicted_mnb_stemmed: predicted values
        :return: confusion matrix with positives and negatives
        """
        return confusion_matrix(test_column_name[column_name_modified], predicted_mnb_stemmed,
                                labels=target_name_for_classification)

    # scoring each row based on its prediction

    def prediction_rate(self, column_name, column_name_modified, train_column_name,
                        test_column_name, model_name, file_name):
        """
        :param column_name: column trained on
        :param column_name_modified: standard values to train on
        :param train_column_name: column name trained
        :param test_column_name: column name for test
        :param model_name: name of the model selected for prediction
        :param file_name: file name to open the predicted values
        :return:
        """
        predict_model_name = 'Predict_' + model_name
        file_path = os.getcwd() + '/Data Sets/' + column_name + '_Prediction/' + file_name
        column_model_name = pd.read_excel(file_path, sheet_name='Sheet1', encoding='utf-8', na_filter=False,
                                          header=0)
        path_to_open_classification_report = os.getcwd() + '/Data Sets/' + column_name + '_Prediction/' + column_name + '_classification_report_' + column_name + '_' + model_name + '.xlsx'
        selected_value = 'Selected_Prediction'
        column_model_name.insert(loc=1, column=selected_value, value=column_model_name[predict_model_name])
        column_model_name.insert(loc=2, column='Prediction_Percent', value=-1)
        classification_report_scores = pd.DataFrame()
        if os.path.isfile(path_to_open_classification_report):
            classification_report_scores = pd.read_excel(path_to_open_classification_report, sheet_name='Sheet1',
                                                         encoding='utf-8', na_filter=False, header=0)
            column_model_name.insert(loc=3, column='F1_Score', value=-0.01)
        column_model_name = column_model_name.drop([predict_model_name], axis=1)
        for index, words_in_a_row in column_model_name.iterrows():
            if os.path.isfile(path_to_open_classification_report):
                for f1_score_row in classification_report_scores.itertuples():
                    if words_in_a_row.Selected_Prediction.strip() == f1_score_row.Labels.strip():
                        column_model_name.loc[index, 'F1_Score'] = f1_score_row[2]
        return column_model_name


"""this class contains a method that is used for creating a classification report and saving it in an excel sheet"""


class ClassificationReports:
    def classifcation_report_processing(self, column_name, model_name, model_to_report):
        """
        :param column_name: column to process
        :param model_name: model name to create classification report
        :param model_to_report: classification report of the selected model
        :return: message if it is saved or not
        """
        message = 'Classification Report Saved'
        tmp = list()
        """splitting the report on new line character"""
        for row in model_to_report.split("\n"):
            parsed_row = [x for x in row.split("  ") if len(x) > 0]
            if len(parsed_row) > 0:
                tmp.append(parsed_row)

        dataframe = pd.DataFrame.from_records(tmp)
        new_header = dataframe.iloc[0]  # grab the first row for the header
        dataframe = dataframe[1:]  # take the data less the header row
        dataframe.columns = new_header
        list_headers = list(dataframe.columns.values)
        list_headers.insert(0, 'Labels')
        list_headers = list_headers[:-1]
        dataframe.columns = list_headers

        path_to_save = os.getcwd() + '/Data Sets/' + column_name + '_Prediction/' + column_name \
                       + '_classification_report_' + column_name + '_' + model_name + '.xlsx'
        dataframe.to_excel(path_to_save, index=False)
        return message


"""This class contains the method which creates user interface and gets the input from user interface for further 
processing"""

class EditableListCtrl(wx.ListCtrl, listmix.TextEditMixin):
    def __init__(self, parent, ID=wx.ID_ANY, pos=wx.DefaultPosition,
                 size=wx.DefaultSize, style=0):

        wx.ListCtrl.__init__(self, parent, ID, pos, size, style)
        listmix.TextEditMixin.__init__(self)


class MainPanel(wx.Panel):
    def __init__(self, parent):
        super().__init__(parent)
        main_sizer = wx.BoxSizer(wx.VERTICAL)
        self.row_obj_dict = {}
        self.current_folder_path = None
        self.list_ctrl = EditableListCtrl(self, size=(600, 800), style=wx.LC_REPORT)

        self.list_ctrl.InsertColumn(0, "WD Customer", width=200)
        self.list_ctrl.InsertColumn(1, "Selected Prediction", width=200)
        self.list_ctrl.InsertColumn(2, "F1 Score", width=200)
        main_sizer.Add(self.list_ctrl, 0, wx.ALL | wx.EXPAND, 5)
        save_button = wx.Button(self, label="Save")
        save_button.Bind(wx.EVT_BUTTON, self.on_save)
        btn_sizer = wx.BoxSizer(wx.HORIZONTAL)

        main_sizer.Add(btn_sizer, 0, wx.ALL | wx.CENTER, 5)

        btn_sizer.Add(save_button, 0, wx.ALIGN_CENTER | 5, 5)
        self.SetSizer(main_sizer)
        self.list_ctrl.Bind(wx.EVT_LIST_END_LABEL_EDIT, self.on_update)

    def on_update(self, event):
        self.list_ctrl.Select(event.Item.Id)  # force the list to select the event item
        row_id = event.GetIndex()  # Get the current row
        col_id = event.GetColumn()  # Get the current column
        if col_id < 0:  # ---- Changed ------
            col_id = 0  # ---- Changed ------
        new_data = event.GetText()  # Get the changed data
        # print(new_data)
        cols = self.list_ctrl.GetColumnCount()  # Get the total number of columns
        rows = self.list_ctrl.GetItemCount()  # Get the total number of rows

        # Get the changed item use the row_id and iterate over the columns
        # print(" ".join([self.list_ctrl.GetItem(row_id, colu_id).GetText() for colu_id in range(cols)]))
        # print("Changed Item:", new_data, "Column:", col_id)

        # Get the entire listctrl iterate over the rows and the columns within each row
        '''print("\nEntire listctrl BEFORE the update:")
        for row in range(rows):
            row_data = (" ".join([self.list_ctrl.GetItem(row, col).GetText() for col in range(cols)]))
            print(row_data)'''

        # Set the new data in the listctrl
        self.list_ctrl.SetStringItem(row_id, col_id, new_data)

        # print("\nEntire listctrl AFTER the update:")
        # Create a list that can be used to export data to a file
        global data_from_listctrl
        data_for_export = []
        for row in range(rows):
            row_data = (";".join([self.list_ctrl.GetItem(row, col).GetText() for col in range(cols)]))
            data_for_export.append(row_data)  # Add to the exportable data

        # print("\nData for export")
        '''for row in data_for_export:  # Print the data
            print(row)'''
        data_from_listctrl = data_for_export

    def on_save(self, event):
        self.savingtofile()

    def savingtofile(self):
        global data_from_listctrl, original_user_frame, original_file_path
        save_path = original_file_path.rsplit('\\', 1)[0]
        save_path = save_path + '\\Final_File.xlsx'

        global input_frame_path
        initial_frame = pd.read_excel(input_frame_path, encoding='utf-8', header=0)
        initial_frame['Initial_Value_Copy'] = [str(word).lower() for word in initial_frame['WD Customer']]
        initial_frame = initial_frame.drop(initial_frame[(initial_frame['Initial_Value_Copy'] == '')].index)
        initial_frame = initial_frame.drop(initial_frame[(initial_frame['Initial_Value_Copy'] == ' ')].index)
        initial_frame = initial_frame.drop(initial_frame[(initial_frame['Initial_Value_Copy'] == 'no_value')].index)

        if data_from_listctrl:
            classfied_frame = pd.DataFrame(data_from_listctrl)
            list_headers = list(classfied_frame.columns.values)
            list_headers.insert(0, 'Initial_Value')
            list_headers = list_headers[:-1]
            classfied_frame.columns = list_headers
            classfied_frame.insert(loc=1, column='Selected_Prediction', value='NA')
            classfied_frame.insert(loc=2, column='F1_Score', value='NA')

            for index, row in classfied_frame.iterrows():
                if ';' in row['Initial_Value']:
                    initial_value, selected_prediction, score = row['Initial_Value'].split(';')
                    classfied_frame.loc[index, 'Initial_Value'] = initial_value
                    classfied_frame.loc[index, 'Selected_Prediction'] = selected_prediction
                    classfied_frame.loc[index, 'F1_Score'] = score

            classfied_frame['Initial_Value_Copy'] = [word.lower() for word in classfied_frame['Initial_Value']]
            classfied_frame = classfied_frame.drop(classfied_frame[(classfied_frame['Initial_Value_Copy'] == '')].index)
            classfied_frame = classfied_frame.drop(classfied_frame[(classfied_frame['Initial_Value_Copy'] ==
                                                                    ' ')].index)
            classfied_frame = classfied_frame.drop(classfied_frame[(classfied_frame['Initial_Value_Copy'] ==
                                                                    'no_value')].index)

            initial_frame.reset_index(drop=True)

            for initial_index, initial_row in initial_frame.iterrows():
                for classified_index, classified_row in classfied_frame.iterrows():
                    if initial_row.Initial_Value_Copy == classified_row.Initial_Value_Copy:
                        initial_frame.loc[initial_index, 'Customer (Aggre)'] = classified_row['Selected_Prediction']

            del initial_frame['Initial_Value_Copy']

            aggrid = initial_frame['Customer (Aggre)'].nunique()
            unique_customer = list(initial_frame['Customer (Aggre)'].unique())

            user_id_dictionary = dict()

            for ids in range(1, aggrid + 1):
                for customer in unique_customer:
                    if customer in user_id_dictionary:
                        pass
                    else:
                        user_id_dictionary[customer] = ids
                        break

            for index, rows in initial_frame.iterrows():
                for key, value in user_id_dictionary.items():

                    if key == rows['Customer (Aggre)']:
                        initial_frame.loc[index, 'Customer_AGRR_ID'] = value

            initial_frame.to_excel(save_path, header=True, index=False)
        else:
            original_user_frame['Initial_Value_Copy'] = [word.lower() for word in original_user_frame['Initial_Value']]
            original_user_frame = \
                original_user_frame.drop(original_user_frame[(original_user_frame['Initial_Value_Copy'] == '')].index)
            original_user_frame = \
                original_user_frame.drop(original_user_frame[(original_user_frame['Initial_Value_Copy'] == ' ')].index)
            original_user_frame = \
                original_user_frame.drop(original_user_frame[(original_user_frame['Initial_Value_Copy'] ==
                                                              'no_value')].index)

            for initial_index, initial_row in initial_frame.iterrows():
                for original_index, original_row in original_user_frame.iterrows():
                    if initial_row.Initial_Value_Copy == original_row.Initial_Value_Copy:
                        initial_frame.loc[initial_index, 'Customer (Aggre)'] = original_row['Selected_Prediction']
            del original_user_frame['Initial_Value_Copy']

            aggrid = initial_frame['Customer (Aggre)'].count()
            unique_customer = list(initial_frame['Customer (Aggre)'].unique())
            user_id_dictionary = dict()
            for ids in range(1, aggrid + 1):
                for customer in unique_customer:
                    user_id_dictionary.update({customer: ids})

            print(user_id_dictionary)
            for index, rows in initial_frame.iterrows():
                for key, value in user_id_dictionary.items():
                    if key == rows['Customer (Aggre)']:
                        initial_frame.loc[index, 'Customer_AGRR_ID'] = value
            initial_frame.to_excel(save_path, header=True, index=False)

    def update_user_listing(self, file_path):
        self.list_ctrl.ClearAll()
        global original_file_path
        original_file_path = file_path
        self.list_ctrl.InsertColumn(0, "WD Customer", width=200)
        self.list_ctrl.InsertColumn(1, "Selected Prediction", width=200)
        self.list_ctrl.InsertColumn(2, "F1 Score", width=200)

        userframe = self.exceloperations()

        user_objects = []
        indexes = 0
        for indexs, users in userframe.iterrows():
            self.list_ctrl.InsertItem(indexes, users.Initial_Value)
            self.list_ctrl.SetItem(indexes, 1, users.Selected_Prediction)
            self.list_ctrl.SetItem(indexes, 2, users.F1_Score)
            user_objects.append(users)
            self.row_obj_dict[indexs] = users
            indexes += 1

    def exceloperations(self):
        file_path = original_file_path
        userframe = pd.read_excel(file_path, encoding='utf-8', header=0, dtype=str)
        global original_user_frame
        original_user_frame = userframe.copy()
        return original_user_frame


class MainFrame(wx.Frame):
    def __init__(self):
        super().__init__(parent=None, size=(600, 900), title="Customer Name Classification")
        icon_path = os.getcwd() + "\\Siemens.png"
        self.SetIcon(wx.Icon(icon_path))
        self.panel = MainPanel(self)
        self.SetBackgroundColour('Grey')
        font = wx.Font(18, wx.DECORATIVE, wx.ITALIC, wx.NORMAL)
        self.SetFont(font)
        self.create_menu()
        self.Show()

    def create_menu(self):
        menu_bar = wx.MenuBar()
        file_menu = wx.Menu()
        open_file_menu_item = file_menu.Append(
            wx.ID_ANY, "Open File", "Open a file to classify customer names"
        )
        menu_bar.Append(file_menu, "&File")
        self.Bind(
            event=wx.EVT_MENU,
            handler=self.on_open_file,
            source=open_file_menu_item,
        )
        self.SetMenuBar(menu_bar)

    def on_open_file(self, event):
        title = "Choose a file:"

        dlg = wx.FileDialog(self, title, style=wx.DD_DEFAULT_STYLE)
        if dlg.ShowModal() == wx.ID_OK:
            """main method that calls the method to create ui and do the rest of the processing"""
            object_process_all_text_columns = ProcessAllTextColumns()
            combined_all_column_output_path = object_process_all_text_columns.ml_process_general_text_columns_processing(dlg.GetPath())
            self.panel.update_user_listing(combined_all_column_output_path)
        dlg.Destroy()


''' Class to run the process of text processing by taking all the inputs and carry out human interaction with machine
 The method to run all the functions of machine learning and non-machine learning process'''


class ProcessAllTextColumns:
    def ml_process_general_text_columns_processing(self, filepath):
        """
        :return: na
        """

        mapped_file = os.getcwd() + '\\Data Sets\\Mapping_Technical_Data.xlsx'

        columns_array = 'Initial_Value'
        input_file_name = filepath
        global input_frame_path
        input_frame_path = filepath
        """If the training file already exists then train the training file using the predicted vlaues"""
        import_train_file_mapping = UniqueValuesToTrainingFile()
        import_train_file_mapping.train_the_training_file(columns_array, input_file_name)
        input_file_path = '0'
        while input_file_path == "0":
            print(input_file_name)
            input_data = pd.DataFrame()
            """checking if the the file path entered by the user exists or not"""
            if os.path.isfile(input_file_name):
                input_file_name_exists = '0'
                while input_file_name_exists == '0':
                    """checking for file if it is csv or xlsx file to read else error message printed as format 
                    not supported"""
                    if input_file_name.rpartition('.')[2] == 'pdf':
                        with open(input_file_name, 'rb') as f:
                            pages = PdfFileReader(f).getNumPages()

                        tables = camelot.read_pdf(input_file_name ,pages='all')
                        # print(tables)
                        input_data = pd.DataFrame()
                        dataframe_page = pd.DataFrame
                        # input_data.rename(columns={list(input_data)[0]: columns_array}, inplace=True)
                        page = 0
                        for page in range(pages):
                            if page == 0:
                                print('Pages: ', pages)
                                input_data = tables[page].df
                            else:
                                dataframe_page = tables[page].df
                                input_data.rename(columns={list(input_data)[0]: columns_array}, inplace=True)
                                dataframe_page.rename(columns={list(dataframe_page)[0]: columns_array},
                                                      inplace=True)

                                if dataframe_page[columns_array].equals(input_data[columns_array]):
                                    dataframe_page = dataframe_page.drop(columns=columns_array, axis=1)
                                    dataframe_page = dataframe_page.drop(columns=1, axis=1)
                                    input_data = pd.concat([input_data, dataframe_page], axis=1)
                                else:
                                    input_data = pd.concat([input_data, dataframe_page], axis=0)

                        input_file_name_exists = '1'
                        if pages == page:
                            break
                        print('Saving PDF as excel for further cleaning.')
                        path_to_save = os.getcwd() + '/Data Sets/Pre_Cleaned_File.xlsx'
                        input_data.to_excel(path_to_save, index=False)
                        doClean = DoCleanInput()
                        input_data = doClean.do_cleaning(mapped_file, path_to_save, 'Sheet1')
                    else:
                        # input_file_sheet_name = args.input_file_sheet_name
                        input_file_sheet_name = 'Tabelle1'
                        # doTrans = DoTranspose()
                        # input_data = doTrans.do_transposer(mapped_file, input_file_name, input_file_sheet_name)
                        # if flag_trans == 1:
                        #     print('Input file needed transpose, transposed and saved, processing further.')
                        #     path_to_open = os.getcwd() + '/Data Sets/Transposed_File.xlsx'
                        #     input_data = pd.read_xlsx(path_to_open, encoding='utf-8', header=0, na_filter=False)

                        if input_data.empty:
                            print('Input file dont need to be transformed, processing further.')
                            dotrans = DoCleanInput()
                            input_data = dotrans.do_cleaning(mapped_file, input_file_name, input_file_sheet_name)

                            # input_data = pd.read_csv(input_file_name, encoding='utf-8', header=0, na_filter=False,
                            #                          error_bad_lines=False)
                            # input_data.rename(columns={list(input_data)[0]: columns_array}, inplace=True)
                        # logging.info('Loading Input file')
                        print(colored('Loading Input file', 'green'))
                        input_file_name_exists = '1'

                        # input_file_name_exists_message = args.input_file_name_exists_message
                        # if (input_file_name_exists_message == '0'):
                        # input_file_path = '0'
                        # else:
                        # break
                # profile = pandas_profiling.ProfileReport(input_data)
                # profile.to_file(os.getcwd() + '\\Data Sets\\' + "Input_File_Profile_Report.html")
                print('Column Coming.. ', columns_array)
                # logging.info('matching Headers with user selected column.')
                header_names = list(input_data.columns.values)
                input_file_path = '1'
                column_name = columns_array
                """getting the list of header from input file and iterating over them"""
                for header in header_names:
                    header_with_underscore = str(header).replace(" ", "_")
                    """if column name given by user exists in input file then do further processing"""
                    if header == column_name or header_with_underscore == column_name:
                        # logging.info('Column Found')
                        directory_to_save_files = header
                        directory_to_save_files = directory_to_save_files.replace(" ", "_")
                        column_name = column_name.replace(" ", "_")
                        """Check if the directory exists else create new to save the files"""
                        data_set_path = os.getcwd() + '\\Data Sets\\'
                        if os.path.isdir(data_set_path):
                            # logging.info('Data Sets Directory Exists.')
                            pass
                        else:
                            # logging.info('Directory Created Data Sets.')
                            os.makedirs(data_set_path)
                        path = os.getcwd() + '\\Data Sets\\' + directory_to_save_files + "_Prediction"
                        # check if the directory for each of the file exists
                        """Check for each new column directory else create it"""
                        if os.path.isdir(path):
                            # logging.info('Column Folder Found')
                            pass
                        else:
                            # logging.info('Column Folder Created')
                            os.makedirs(path)
                        print(colored("Directory created, checking for Train Data set...!!!", 'green'))
                        import_text_processing_class = GeneralMethodsClassforTextProcessing()
                        column_name_modified = column_name + '_Modified'
                        """Calling the function to select the unique values from input file"""
                        column_name_head = import_text_processing_class.create_unique_values(input_data,
                                                                                             column_name,
                                                                                             column_name_modified,
                                                                                             header)
                        path_train_file = os.getcwd() + '\\Data Sets\\' + column_name + '_Prediction\\Train_' \
                                          + column_name + '.xlsx'
                        path_test_file = os.getcwd() + '\\Data Sets\\' + column_name + '_Prediction\\Test_' \
                                         + column_name + '.xlsx'
                        print(colored("Train Data set file created...!!!", 'green'))
                        # logging.info('Train Data set file created')

                        picklist_column_file_path = '0'

                        mapped_file_sheet_name = column_name
                        """checking if the label file exists as the input given by user"""
                        if os.path.isfile(mapped_file):
                            check_xl_sheet = pd.ExcelFile(mapped_file)
                            """reading it in either csv or xlsx format"""
                            if mapped_file.rpartition('.')[2] == 'csv':
                                train_column = pd.read_csv(mapped_file, encoding='utf-8',
                                                           header=0, na_filter=False)
                                # logging.info('Train File Created')
                                train_column.columns = [column_name, column_name_modified]
                                if not os.path.exists(os.getcwd() + '/Training_Files'):
                                    os.makedirs(os.getcwd() + '/Training_Files')
                                path_to_save = \
                                    os.getcwd() + '/Data Sets/Training_Files/To_Process_Train_' + column_name + '.xlsx'
                                if not os.path.isfile(path_to_save):
                                    train_column.to_excel(path_to_save, index=False)
                            elif mapped_file.rpartition('.')[2] == 'xlsx':
                                train_column = pd.read_excel(mapped_file,
                                                             encoding='utf-8', na_filter=False, header=0)
                                # logging.info('Train File Created')

                                train_column.columns = [column_name, column_name_modified]
                                if not os.path.exists(os.getcwd() + '/Data Sets/Training_Files'):
                                    os.makedirs(os.getcwd() + '/Data Sets/Training_Files')
                                path_to_save = \
                                    os.getcwd() + '/Data Sets/Training_Files/To_Process_Train_' + column_name + '.xlsx'
                                train_column.to_excel(path_to_save, index=False)
                            elif mapped_file_sheet_name in check_xl_sheet.sheet_names:
                                train_column = pd.read_excel(mapped_file, sheet_name=mapped_file_sheet_name,
                                                             encoding='utf-8', na_filter=False, header=0)
                                # logging.info('Train File Created')
                                train_column.columns = [column_name, column_name_modified]
                                if not os.path.exists(os.getcwd() + '/Data Sets/Training_Files'):
                                    os.makedirs(os.getcwd() + '/Data Sets/Training_Files')
                                path_to_save = \
                                    os.getcwd() + '/Data Sets/Training_Files/To_Process_Train_' + column_name + '.xlsx'
                                if not os.path.isfile(path_to_save):
                                    train_column.to_excel(path_to_save, index=False)
                            else:
                                """else use the piclist file if label file doesnot exist to create the first 
                                train file"""
                                while picklist_column_file_path == '0':
                                    picklist_column_file = os.getcwd() + '\\Data Sets\\Mapping_Technical_Data.xlsx'
                                    if os.path.isfile(picklist_column_file):
                                        picklist_column_file_path = '1'
                                        picklist_column_file_exists = '0'
                                        while picklist_column_file_exists == '0':
                                            if picklist_column_file.rpartition('.')[2] == 'csv':
                                                picklist_column = pd.read_csv(picklist_column_file,
                                                                              encoding='utf-8',
                                                                              header=0, na_filter=False)
                                                picklist_column_file_exists = '1'
                                            elif picklist_column_file.rpartition('.')[2] == 'xlsx':
                                                picklist_column = pd.read_excel(picklist_column_file,
                                                                                sheet_name='Tabelle1',
                                                                                encoding='utf-8', na_filter=False,
                                                                                header=0)
                                                picklist_column_file_exists = '1'
                                            else:
                                                print("file format not supported, please enter the file in .xlsx "
                                                      "or .csv format.")
                                                # logging.info('Train File not found')
                                    else:
                                        print(colored("The file at the give path does not exist.", 'red'))

                                if os.path.isfile(path_train_file):
                                    print("Training File Already Exists using the same.")
                                    # logging.info('Train File Exists')
                        else:
                            while picklist_column_file_path == '0':
                                """if user didnot enter a lable file then use picklist file"""
                                picklist_column_file = os.getcwd() + 'Data Sets\\Mapping_Technical_Data.xlsx'
                                if os.path.isfile(picklist_column_file):
                                    picklist_column_file_path = '1'
                                    picklist_column_file_exists = '0'
                                    while picklist_column_file_exists == '0':
                                        if picklist_column_file.rpartition('.')[2] == 'csv':
                                            picklist_column = pd.read_csv(picklist_column_file, encoding='utf-8',
                                                                          header=0, na_filter=False)
                                            picklist_column_file_exists = '1'
                                            # logging.info('Train File Created')
                                        elif picklist_column_file.rpartition('.')[2] == 'xlsx':
                                            # logging.info('Train File Created')
                                            picklist_column = pd.read_excel(picklist_column_file,
                                                                            sheet_name='Tabelle1',
                                                                            encoding='utf-8', na_filter=False,
                                                                            header=0)
                                            picklist_column_file_exists = '1'
                                else:
                                    print(colored("The file at the give path does not exist.", 'red'))

                            """check if training file already exists else create a new training file"""
                            if os.path.isfile(path_train_file):
                                print("Training File Already Exists using the same.")
                        print(colored('Loading and cross Checking the Train File.', 'green'))
                        """balancing the training file before processing further"""
                        import_text_processing_class.dataset_balancing_in_train_file(
                            column_name, column_name_modified)
                        if os.path.isfile(path_train_file):
                            print(colored("Creating Test file.\nPlease wait progress will be updated below...!!!",
                                          'green'))
                        else:
                            message_train = path_train_file + \
                                            "Files still not found, quitting the current column, " \
                                            "if more column exists, the process will continue."
                            break
                        """create new test file with unique values from input file"""
                        import_text_processing_class.create_test_file(input_data, column_name,
                                                                      column_name_modified, header)

                        """Read Training file and create validation set"""
                        train_column_name = import_text_processing_class.read_train_file(column_name)
                        x_train, x_val, y_train, y_val = train_test_split(train_column_name[column_name],
                                                                          train_column_name[column_name_modified],
                                                                          test_size=0.3, random_state=1)
                        test_column_name = import_text_processing_class.read_test_file(column_name)
                        x_train_tfidf = import_text_processing_class.vectorize_data(column_name, train_column_name)

                        all_column_dataframe = pd.DataFrame()
                        # Models
                        """Run the process through all the models Naive Bayes, SVM, Decision Tree, Random Forest
                        and get the accuracy score for each so that the best model can be selected"""
                        predict_val_nb = import_text_processing_class.validation_nb_model_fit(x_train, y_train,
                                                                                              x_val, x_train_tfidf)
                        predict_nb = import_text_processing_class.nb_model_fit(column_name, column_name_modified,
                                                                               train_column_name, test_column_name,
                                                                               x_train_tfidf)
                        import_text_processing_class.nb_model_predict(column_name, column_name_modified,
                                                                      train_column_name, test_column_name,
                                                                      predict_nb, all_column_dataframe)
                        classification_report_nb = import_text_processing_class.classification_report_nb(y_val,
                                                                                                         predict_val_nb)
                        print("Naive Bayes ", classification_report_nb)
                        predict_val_rf = import_text_processing_class.validation_rf_model_fit(x_train, y_train,
                                                                                              x_val, x_train_tfidf)
                        predict_rf = import_text_processing_class.rf_model_fit(column_name, column_name_modified,
                                                                               train_column_name, test_column_name,
                                                                               x_train_tfidf)
                        import_text_processing_class.rf_model_predict(column_name, column_name_modified,
                                                                      train_column_name, test_column_name,
                                                                      predict_rf, all_column_dataframe)
                        classification_report_rf = import_text_processing_class.classification_report_rf(y_val,
                                                                                                         predict_val_rf)
                        print("Random Forest ", classification_report_rf)
                        predict_val_dt = import_text_processing_class.validation_dt_model_fit(x_train, y_train,
                                                                                              x_val, x_train_tfidf)
                        predict_dt = import_text_processing_class.dt_model_fit(column_name, column_name_modified,
                                                                               train_column_name, test_column_name,
                                                                               x_train_tfidf)
                        import_text_processing_class.dt_model_predict(column_name, column_name_modified,
                                                                      train_column_name, test_column_name,
                                                                      predict_dt, all_column_dataframe)
                        classification_report_dt = import_text_processing_class.classification_report_dt(y_val,
                                                                                                         predict_val_dt)
                        print("Decision Tree ", classification_report_dt)
                        predict_val_svm = import_text_processing_class.validation_svm_model_fit(x_train, y_train,
                                                                                                x_val,
                                                                                                x_train_tfidf)
                        predict_svm = import_text_processing_class.svm_model_fit(column_name, column_name_modified,
                                                                                 train_column_name,
                                                                                 test_column_name,
                                                                                 x_train_tfidf)
                        import_text_processing_class.svm_model_predict(column_name, column_name_modified,
                                                                       train_column_name, test_column_name,
                                                                       predict_svm, all_column_dataframe)
                        classification_report_svm = import_text_processing_class.classification_report_svm(
                            y_val, predict_val_svm)

                        print("SVM ", classification_report_svm)
                        predict_val_svm_gs = import_text_processing_class.validation_svm_grid_search_model_fit(
                            x_train,
                            y_train,
                            x_val,
                            x_train_tfidf)
                        predict_svm_gs = import_text_processing_class.svm_grid_search_model_fit(column_name,
                                                                                                column_name_modified,
                                                                                                train_column_name,
                                                                                                test_column_name,
                                                                                                x_train_tfidf)
                        import_text_processing_class.svm_grid_search_model_predict(column_name, column_name_modified,
                                                                                   train_column_name,
                                                                                   test_column_name,
                                                                                   predict_svm_gs,
                                                                                   all_column_dataframe)
                        classification_report_gs_svm = import_text_processing_class.classification_report_gs_svm(
                            y_val,
                            predict_val_svm_gs)
                        print("SVM GS ", classification_report_gs_svm)
                        predict_val_nb_nltk = import_text_processing_class.validation_nltk_nb_model_fit(x_train,
                                                                                                        y_train,
                                                                                                        x_val,
                                                                                                        x_train_tfidf)
                        predict_nb_nltk = import_text_processing_class.nltk_nb_model_fit(column_name,
                                                                                         column_name_modified,
                                                                                         train_column_name,
                                                                                         test_column_name,
                                                                                         x_train_tfidf)
                        import_text_processing_class.nltk_nb_model_predict(column_name, column_name_modified,
                                                                           train_column_name, test_column_name,
                                                                           predict_nb_nltk, all_column_dataframe)
                        classification_report_nltk_nb = import_text_processing_class.classification_report_nltk_nb(
                            y_val,
                            predict_val_nb_nltk)
                        print("Naive Bayes with NLTK ", classification_report_nltk_nb)
                        """selecting the model with highest accuracy score"""
                        highest = max(classification_report_nb, classification_report_nltk_nb,
                                      classification_report_gs_svm, classification_report_svm,
                                      classification_report_rf, classification_report_dt)
                        import_classification_report_class = ClassificationReports()
                        all_column_dataframe.to_excel(
                            os.getcwd() + '\\Data Sets\\' + column_name + '_Prediction\\Prediction_of_all_models.xlsx',
                            index=False)
                        """Creating Classification reports and confusion matrix for each model outputs"""
                        report_nb = classification_report(y_val, predict_val_nb)
                        matrix_nb = confusion_matrix(y_val, predict_val_nb,
                                                     labels=train_column_name[column_name_modified].unique())
                        print(matrix_nb)
                        saving_cl_report_nb = import_classification_report_class.classifcation_report_processing(
                            column_name, 'NB', report_nb)
                        print(saving_cl_report_nb)
                        report_dt = classification_report(y_val, predict_val_dt)
                        matrix_dt = confusion_matrix(y_val, predict_val_dt,
                                                     labels=train_column_name[column_name_modified].unique())
                        print(matrix_dt)
                        saving_cl_report_dt = import_classification_report_class.classifcation_report_processing(
                            column_name, 'DT', report_dt)
                        print(saving_cl_report_dt)
                        report_rf = classification_report(y_val, predict_val_rf)
                        matrix_rf = confusion_matrix(y_val, predict_val_rf,
                                                     labels=train_column_name[column_name_modified].unique())
                        print(matrix_rf)
                        saving_cl_report_rf = import_classification_report_class.classifcation_report_processing(
                            column_name, 'RF', report_rf)
                        print(saving_cl_report_rf)
                        report_svm = classification_report(y_val, predict_val_svm)
                        matrix_svm = confusion_matrix(y_val, predict_val_svm,
                                                      labels=train_column_name[column_name_modified].unique())
                        print(matrix_svm)
                        saving_cl_report_svm = import_classification_report_class.classifcation_report_processing(
                            column_name, 'SVM', report_svm)
                        print(saving_cl_report_svm)
                        report_svm_gs = classification_report(y_val, predict_val_svm_gs)
                        saving_cl_report_svm_gs = import_classification_report_class.classifcation_report_processing(
                            column_name, 'SVM_GS', report_svm_gs)
                        print(saving_cl_report_svm_gs)
                        matrix_svm_gs = confusion_matrix(y_val, predict_val_svm_gs,
                                                         labels=train_column_name[column_name_modified].unique())
                        print(matrix_svm_gs)
                        report_nltk_nb = classification_report(y_val, predict_val_nb_nltk)
                        matrix_nltk_nb = confusion_matrix(y_val, predict_val_nb_nltk,
                                                          labels=train_column_name[column_name_modified].unique())
                        print(matrix_nltk_nb)
                        saving_cl_report_nb_nltk = import_classification_report_class.classifcation_report_processing(
                            column_name, 'NLTK_NB', report_nltk_nb)
                        print(saving_cl_report_nb_nltk)
                        print("Model with the highest score is selected automatically.")
                        # Rating the prediction with classigication report and confusion matrix from the best model
                        """scoring with string similarity for the highest performing model"""
                        if classification_report_nb == highest:
                            column_model_name = import_text_processing_class.prediction_rate(
                                column_name, column_name_modified, train_column_name, test_column_name, 'NB',
                                'Prediction_of_all_models.xlsx')

                        elif classification_report_dt == highest:
                            column_model_name = import_text_processing_class.prediction_rate(
                                column_name, column_name_modified, train_column_name, test_column_name, 'DT',
                                'Prediction_of_all_models.xlsx')

                        elif classification_report_rf == highest:
                            column_model_name = import_text_processing_class.prediction_rate(
                                column_name, column_name_modified, train_column_name, test_column_name, 'RF',
                                'Prediction_of_all_models.xlsx')

                        elif classification_report_svm == highest:
                            column_model_name = import_text_processing_class.prediction_rate(
                                column_name, column_name_modified, train_column_name, test_column_name, 'SVM',
                                'Prediction_of_all_models.xlsx')

                        elif classification_report_gs_svm == highest:
                            column_model_name = import_text_processing_class.prediction_rate(
                                column_name, column_name_modified, train_column_name, test_column_name, 'SVM_GS',
                                'Prediction_of_all_models.xlsx')

                        elif classification_report_nltk_nb == highest:
                            column_model_name = import_text_processing_class.prediction_rate(
                                column_name, column_name_modified, train_column_name, test_column_name, 'NLTK_NB',
                                'Prediction_of_all_models.xlsx')

                        column_prediction_sorting_order = '0'
                        while column_prediction_sorting_order == '0':
                            """sorting the prediction file before saving based on the user input from best to 
                            worst or vice-versa"""
                            column_prediction_sorting_order = '1'  # args.column_prediction_sorting_order
                            if column_prediction_sorting_order == '1':
                                column_model_name = column_model_name.sort_values('F1_Score')
                            elif column_prediction_sorting_order == '2':
                                column_model_name = column_model_name.sort_values('F1_Score', ascending=False)
                            else:
                                column_prediction_sorting_message = '1'  # args.column_prediction_sorting_message
                                if column_prediction_sorting_message == '0':
                                    column_prediction_sorting_order = '0'
                                else:
                                    column_prediction_sorting_order = '3'
                                    print('Choice not found, file will be saved without sorting.')
                        file_to_save = pd.DataFrame()
                        column_model_name = column_model_name.drop(['Prediction_Percent'], axis=1)
                        file_to_save = column_model_name.copy()
                        """Creating new directory to save all the predictions combined"""
                        input_file_name_folder_for_columns = input_file_name.rsplit('\\', 1)[1]
                        path_to_save = \
                            os.getcwd() + '\\Data Sets\\' + column_name + '_Prediction\\Predicted_Score_' \
                            + column_name + '.xlsx'
                        if not os.path.exists(os.getcwd() + '/Data Sets/Combined_all_column_output'):
                            os.makedirs(os.getcwd() + '/Data Sets/Combined_all_column_output')
                        if not os.path.exists(
                                os.getcwd() + '/Data Sets/Combined_all_column_output/' +
                                input_file_name_folder_for_columns):
                            os.makedirs(
                                os.getcwd() + '/Data Sets/Combined_all_column_output/' +
                                input_file_name_folder_for_columns)
                        global path_to_save_combined_all_column_output
                        path_to_save_combined_all_column_output = \
                            os.getcwd() + '\\Data Sets\\Combined_all_column_output\\' + \
                            input_file_name_folder_for_columns + '\\Predicted_Score_all_columns.xlsx'

                        save_combined_all_column_output = column_model_name.copy()
                        cleanedframefilepath = os.getcwd() + '\\Data Sets\\Cleaned_File.xlsx'

                        saver = Save_With_All_Values()
                        saver.saving_all_values(save_combined_all_column_output, cleanedframefilepath,
                                                input_file_name_folder_for_columns)

                        """Saving all the predicted models together in one file"""
                        if not os.path.exists(path_to_save_combined_all_column_output):
                            with pd.ExcelWriter(path_to_save_combined_all_column_output,
                                                engine='xlsxwriter') as writer:
                                save_combined_all_column_output.to_excel(writer, column_name, index=False)
                                writer.save()
                        else:
                            book = load_workbook(path_to_save_combined_all_column_output)
                            with pd.ExcelWriter(path_to_save_combined_all_column_output,
                                                engine='openpyxl') as writer:
                                writer.book = book
                                writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
                                save_combined_all_column_output.to_excel(writer, column_name, index=False)
                                writer.save()

                        file_to_save.to_excel(path_to_save, index=False)
                        file_to_save.head()
                print(colored("The process finished for the give column.", 'green'))
            else:
                print(colored("The file at the give path does not exist.", 'red'))
                input_file_path_message_colored = colored(
                    "Do you wish to try again.\nPress 0 to try again or any key to exit\nYour Response: ", 'blue')
                input_file_path_message = 1
                if input_file_path_message == '0':
                    input_file_path = '0'
                else:
                    break

            response = 2
            if response == '1':
                input_file_path = '0'
            else:
                pass
        return path_to_save_combined_all_column_output


"""main method to run the Text processing"""

def main():
    app = wx.App(False)
    frame = MainFrame()
    app.MainLoop()


if __name__ == '__main__':
    main()
