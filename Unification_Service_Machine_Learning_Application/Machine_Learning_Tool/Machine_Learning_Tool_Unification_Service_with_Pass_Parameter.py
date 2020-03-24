# coding: utf-8
# from Save_as_JSON import Save_as_JSON
from SavewithAllValues import Save_With_All_Values
from New_File_Cleaning import DoCleanInput
# from ldaProjectTransposeMechanism import DoTranspose
import datetime
#import logging
import inspect, os
import re
import string
import sys
from collections import defaultdict

import numpy as np
import pandas as pd

from gooey import Gooey, GooeyParser
from nltk.stem import PorterStemmer
from nltk.stem.snowball import SnowballStemmer
from openpyxl import load_workbook

from sklearn.ensemble import RandomForestClassifier
from sklearn.feature_extraction.text import TfidfVectorizer, CountVectorizer, TfidfTransformer
from sklearn.linear_model import SGDClassifier
from sklearn.metrics import accuracy_score, f1_score
from sklearn.metrics import classification_report
from sklearn.metrics import confusion_matrix
from sklearn.model_selection import GridSearchCV
from sklearn.model_selection import train_test_split
from sklearn.naive_bayes import MultinomialNB
from sklearn.pipeline import Pipeline
from sklearn.tree import DecisionTreeClassifier
from sklearn.utils import shuffle
from termcolor import colored
from textblob import TextBlob
from textblob import Word
import camelot
from PyPDF2 import PdfFileReader

st = PorterStemmer()
"""
This class updated all the new values from the prediction file to the training file.
This will work from the second run as there are no available predictions during the first run.
<Excel file name> 
<Excel file sheet name to process>
"""


class Unique_Values_To_Training_File:
    def Training_The_Train_Main_Method(self, file, column_name, column_name_modified, sheet_name_use):
        """
        :param file: path for the prediction file
        :param column_name: column to update in training file
        :param column_name_modified: column's lable values to be updated
        :param sheet_name_use: sheet name for the predicted file
        :return: message if the file is update or not
        """
        self.column_name = column_name
        self.col_name_modified = column_name_modified
        self.file = file
        self.sheet_name_use = sheet_name_use
        path_to_open_predicted_file = file
        """opening an excel fill and returing a read only file inorder to get the list of columns exisitng"""
        wb = load_workbook(path_to_open_predicted_file, read_only=True)  # open an Excel file and return a workbook
        """checking if column exists in the predicted file
        """
        if column_name in wb.sheetnames:
            path_to_open_training_file = os.getcwd() + '/Data Sets/Training_Files/To_Process_Train_' + column_name + '.xlsx'
            """reading the traning file"""
            training_file = pd.read_excel(path_to_open_training_file, sheet_name='Sheet1', encoding='utf-8',
                                          na_filter=False,
                                          header=0)
            training_file[column_name] = training_file[column_name].map(lambda x: str(x).strip())
            training_file[column_name] = [word.lower() for word in training_file[column_name]]
            """reading the prediction file"""
            predicted_file = pd.read_excel(path_to_open_predicted_file, sheet_name=sheet_name_use, encoding='utf-8',
                                           na_filter=False,
                                           header=0)
            predicted_file[column_name] = predicted_file[column_name].map(lambda x: x.strip())
            predicted_file[column_name] = [word.lower() for word in predicted_file[column_name]]

            """iterating over the prediction and the training dataframes"""
            for indexes_output_file, rows_predicted_file in predicted_file.iterrows():
                flag = 1
                for indexes_input_file, rows_training_file in training_file.iterrows():
                    """cross checking if there are new values in the prediction file
                    this is done by checking if the each row value exists in both the dataframes or not
                    if yes setting the flag to 0. so that we dont update the new values to training dataframe.
                    else we do update the values"""
                    if rows_predicted_file[column_name] == rows_training_file[column_name]:
                        flag = 0
                        break
                """adding the new values at the end of training dataframe"""
                if (flag == 1):
                    training_file = training_file.append({column_name: rows_predicted_file[column_name],
                                                          column_name_modified: rows_predicted_file[
                                                              'Selected_Prediction']}, ignore_index=True)

            """iterating over both the dataframes again 
            but this time check if the values are changed in prediction file"""
            for indexes_input_file, rows_predicted_file in predicted_file.iterrows():
                for indexes_output_file, rows_training_file in training_file.iterrows():

                    """if the column values exists in both the files
                    and if the predicted value is not same as the training dataframe label value
                    then we update the new value to it."""
                    if rows_predicted_file[column_name] == rows_training_file[column_name]:
                        if rows_predicted_file['Selected_Prediction'] != rows_training_file[column_name_modified]:
                            training_file.loc[indexes_output_file, column_name_modified] = rows_predicted_file[
                                'Selected_Prediction']

            path_to_save = os.getcwd() + '/Data Sets/Training_Files/To_Process_Train_' + column_name + '.xlsx'
            """Saving the changes to training file"""
            training_file.to_excel(path_to_save, index=False)
            message = 'Train File Updated'
        else:
            message = 'Train File not updated'

        return message

    def Train_The_Training_File(self, column_name, input_file_name):
        """
        :param column_name: column to update in training file
        :param input_file_name: input file with path
        :return: message if both file exists or not. if exists then if it is updated or not
        """
        self.column_name = column_name
        self.input_file_name = input_file_name
        """Getting the folder name where the files are located"""
        input_file_folder_name = input_file_name
        print('input_file_folder_name', input_file_folder_name)
        """calling the constructor of traing train file class"""
        train_the_training = Unique_Values_To_Training_File()
        column_name = column_name.replace(" ", "_")
        column_name_modified = column_name + '_Modified'
        """checking if the training file exists or not
        and also if the prediction file exists or not
        if both exists then we call the method to update the training file"""
        if (os.path.isfile(os.getcwd() + '/Data Sets/Training_Files/To_Process_Train_' + column_name + '.xlsx') and
                os.path.isfile(
                    os.getcwd() + '/Data Sets/Combined_all_column_output/' + input_file_folder_name + '/Predicted_Score_all_columns.xlsx')):
            file = os.getcwd() + '/Data Sets/Combined_all_column_output/' + input_file_folder_name + '/Predicted_Score_all_columns.xlsx'
            message = train_the_training.Training_The_Train_Main_Method(file, column_name, column_name_modified,
                                                                        column_name)
        elif (os.path.isfile(os.getcwd() + '/Data Sets/Training_Files/To_Process_Train_' + column_name + '.xlsx') and
              os.path.isfile(
                  os.getcwd() + '/Data Sets/' + column_name + '_Prediction/Predicted_Score_' + column_name + '.xlsx')):
            file = os.getcwd() + '/Data Sets/' + column_name + '_Prediction/Predicted_Score_' + column_name + '.xlsx'
            message = train_the_training.Training_The_Train_Main_Method(file, column_name, column_name_modified,
                                                                        'Sheet1')
        else:
            message = 'Train File or Prediction File Does not exist'
        return message



# all the methods to do text processing are under the below class
"""This calss holds all the methods that are necessary for text processing.
This class is directly called from the main class that checks all the file and then call each method seperately
depending on the need of method."""


class General_Methods_Class_for_Text_Processing:
    # method to create the unique values file and then manually map them to picklist
    def create_unique_values(self, input_data, column_name, column_name_modified, header):
        """
        :param input_data: input file datarame
        :param column_name: column to process
        :param column_name_modified: column with standard values
        :param header: header of the column
        :return:
        """
        self.input_data = input_data
        self.column_name = column_name
        self.column_name_modified = column_name_modified
        self.header = header
        """creating empty dataframe with headers"""
        column_name_train = pd.DataFrame(columns=[column_name, column_name_modified])
        """adding the values to column from the input dataframe"""
        column_name_train[column_name] = input_data[header]
        """removing empty values"""
        column_name_train = column_name_train.drop(column_name_train[(column_name_train[column_name] == '')].index)
        """removing all the duplicate values"""
        column_name_train.drop_duplicates(subset=[column_name], keep='first', inplace=True)

        path_to_save = os.getcwd() + '/Data Sets/' + column_name + '_Prediction/Picklist_Train_' + column_name + '.xlsx'
        """saving the unique values into a new file for further processing in next steps"""
        column_name_train.to_excel(path_to_save, index=False)
        return column_name_train.head()

    # to do automatic mapping of train file with pick list

    def map_train_file_with_picklist(self, column_name, column_name_modified, header, picklist_column):
        """
        :param column_name: column to map with picklist
        :param column_name_modified: standard values for each column
        :param header: header of the column
        :param picklist_column: column from the picklist table
        :return:
        """
        self.column_name = column_name
        self.column_name_modified = column_name_modified
        self.header = header
        self.picklist_column = picklist_column

        """removing unwanted characters from the picklist dataframe headers"""
        picklist_column.columns = [x.strip() for x in picklist_column.columns]
        column_name_copy = column_name + '_Copy'
        """opening a the train file created in last step"""
        path_to_open_train_file = os.getcwd() + '/Data Sets/' + column_name + '_Prediction/Picklist_Train_' + column_name + '.xlsx'
        train_column_name = pd.read_excel(path_to_open_train_file, sheet_name='Sheet1', encoding='utf-8',
                                          na_filter=False, header=0)
        """creating a new empty dataframe with headers"""
        train_column_name_with_picklist = pd.DataFrame(columns=[column_name, column_name_modified, column_name_copy])
        """adding the exisitng train dataframe values to all three columns"""
        train_column_name_with_picklist[column_name] = train_column_name[column_name]
        train_column_name_with_picklist[column_name_modified] = train_column_name[column_name_modified]
        train_column_name_with_picklist[column_name_copy] = train_column_name[column_name]
        """striping spaces"""
        train_column_name_with_picklist[column_name] = train_column_name_with_picklist[column_name].map(
            lambda x: x.strip())
        """dropping empty values"""
        train_column_name_with_picklist = train_column_name_with_picklist.drop(
            train_column_name_with_picklist[(train_column_name_with_picklist[column_name_copy] == '')].index)
        """dropping rows with just space as value"""
        train_column_name_with_picklist = train_column_name_with_picklist.drop(
            train_column_name_with_picklist[(train_column_name_with_picklist[column_name_copy] == ' ')].index)
        """converting to lower case"""
        train_column_name_with_picklist[column_name_copy] = [word.lower() for word in
                                                             train_column_name_with_picklist[column_name_copy]]
        """replacing / with space"""
        train_column_name_with_picklist[column_name_copy] = train_column_name_with_picklist[
            column_name_copy].str.replace('/', ' ')
        """replacing ; with space"""
        train_column_name_with_picklist[column_name_copy] = train_column_name_with_picklist[
            column_name_copy].str.replace(';', ' ')
        """replacing , with space"""
        train_column_name_with_picklist[column_name_copy] = train_column_name_with_picklist[
            column_name_copy].str.replace(',', ' ')
        """replacing alpha numeric and white spaces with single space"""
        train_column_name_with_picklist[column_name_copy] = train_column_name_with_picklist[
            column_name_copy].str.replace('[^\w\s]', ' ')
        """stemming the words"""
        train_column_name_with_picklist[column_name_copy] = train_column_name_with_picklist[column_name_copy].apply(
            lambda x: " ".join([st.stem(word) for word in x.split()]))

        column_name_modified = column_name + '_Modified'
        """creating another dataframe with headers"""
        picklist_column_name = pd.DataFrame(columns=[column_name, column_name_modified])
        picklist_column_name[column_name] = picklist_column[header]
        picklist_column_name[column_name_modified] = picklist_column[header.strip()]

        """processing picklist through the same steps"""
        picklist_column_name = picklist_column_name.drop(
            picklist_column[(picklist_column_name[column_name_modified] == '')].index)
        picklist_column_name[column_name_modified] = picklist_column_name[column_name_modified].map(lambda x: x.strip())
        picklist_column_name[column_name_modified] = [word.lower() for word in
                                                      picklist_column_name[column_name_modified]]
        picklist_column_name[column_name_modified] = picklist_column_name[column_name_modified].str.replace('[^\w\s]',
                                                                                                            ' ')
        picklist_column_name[column_name_modified] = picklist_column_name[column_name_modified].apply(
            lambda x: " ".join([st.stem(word) for word in x.split()]))

        """Here we will do the string similarity
        iterating over the training file and the picklist file after preprocessing steps"""
        for words in train_column_name_with_picklist.itertuples():
            for word in words[3].split():
                for picklist_words in picklist_column_name.itertuples():
                    for picklist_word in picklist_words[2].split():
                        texx = re.search(r"\b{}\b".format(picklist_word), word, re.IGNORECASE) is not None
                        """if the word matches then we add it to the standard values"""
                        if texx:
                            train_column_name_with_picklist.set_value(words.Index, column_name_modified,
                                                                      picklist_words[1])
        """marking the filled and empty values next to each row 
        so that the user can add values to empty values in excel file"""
        for words in train_column_name_with_picklist.itertuples():
            if words[2] == '':
                train_column_name_with_picklist.set_value(words.Index, 'Filled', 'False')
            else:
                train_column_name_with_picklist.set_value(words.Index, 'Filled', 'True')

        """creating a new dataframe with headers and saving the new values in it"""
        save_train_data = pd.DataFrame(columns=[column_name, column_name_modified, 'Filled'])
        save_train_data[column_name] = train_column_name_with_picklist[column_name]
        save_train_data[column_name_modified] = train_column_name_with_picklist[column_name_modified]
        save_train_data['Filled'] = train_column_name_with_picklist['Filled']
        save_train_data.head(20)

        """checking if training directory exists if not then creating a new directory to save the file"""
        if not os.path.exists(os.getcwd() + '/Data Sets/Training_Files'):
            os.makedirs(os.getcwd() + '/Data Sets/Training_Files')
        path_to_save = os.getcwd() + '/Data Sets/Training_Files/To_Process_Train_' + column_name + '.xlsx'
        save_train_data.to_excel(path_to_save, index=False)
        return train_column_name.head()

    # Normalize the count of all the picklist items in the data for more accurate prediction

    def dataset_balancing_in_train_file(self, column_name, column_name_modified, header):
        """
        :param column_name: Column to process
        :param column_name_modified: Standard values
        :param header: Header of column
        :return: First 5 values of the dataframe
        """
        self.column_name = column_name
        self.column_name_modified = column_name_modified
        self.header = header
        """reading the train file"""
        path_to_open = os.getcwd() + '/Data Sets/Training_Files/To_Process_Train_' + column_name + '.xlsx'
        train_column_name = pd.read_excel(path_to_open, sheet_name='Sheet1', encoding='utf-8', na_filter=False,
                                          header=0)
        """Setting the value in dataframe if it standard value is empty"""
        for index, words in train_column_name.iterrows():
            if words[column_name_modified] == '':
                train_column_name.loc[index, 'Filled'] = 'False'
                #train_column_name.set_value(words.Index, 'Filled', 'False')
            else:
                train_column_name.loc[index, 'Filled'] = 'True'
                #train_column_name.set_value(words.Index, 'Filled', 'True')

        column_name_used = column_name + '_Used'
        """Grouping by column and the unique values"""
        df = train_column_name.groupby(column_name_modified)[column_name].nunique()
        df_highest = df.nlargest(1)

        df_list = []
        """basic preprocessing"""
        train_column_name[column_name_used] = train_column_name[column_name_modified]
        train_column_name[column_name_used] = train_column_name[column_name_used].str.replace('/', '_')
        train_column_name[column_name_used] = train_column_name[column_name_used].str.replace(';', '_')
        train_column_name[column_name_used] = train_column_name[column_name_used].str.replace(',', '_')
        train_column_name[column_name_used] = train_column_name[column_name_used].str.replace(" ", "")

        count = 0
        """Counting the occurence of each standard values and storing their count in a list"""
        for row in train_column_name.itertuples():
            for i, v in df.iteritems():
                if v <= df_highest.values:
                    if row[2] == i:
                        df_column = 'df_' + str(row[4])
                        if not df_list:
                            count += 1
                            df_list.append(df_column)
                        elif df_column not in df_list:
                            count += 1
                            df_list.append(df_column)
                        else:
                            pass
        """Sorting the training dataframe with standard values"""
        train_column_name_sorted = train_column_name.sort_values(column_name_used)
        df_processed = pd.DataFrame()

        count_df = 0
        """Iterating over the list with the count of each standard value"""
        for items in df_list:
            count_df += 1
            items_df = 'df_' + str(count_df)
            items_df = pd.DataFrame()
            count = 0
            """iterating over the sorted dataframe"""
            for index, row in train_column_name_sorted.iterrows():
                df_val = 'df_' + row[column_name_used]
                """Adding values to different dataframes for each standard values"""
                if df_val == items:
                    items_df.loc[count, column_name] = row[column_name]
                    items_df.loc[count, column_name_modified] = row[column_name_modified]
                    items_df.loc[count, column_name_used] = row[column_name_used]

                    # items_df.set_value(count, column_name, row[1])
                    # items_df.set_value(count, column_name_modified, row[2])
                    # items_df.set_value(count, column_name_used, row[4])
                    count += 1
            """running while all the elements in the dataframe are not read"""
            while (len(items_df.index) <= df_highest[0]):
                """Running through different conditions to balance the dataset"""
                if len(items_df.index) <= (df_highest[0] / 2):
                    item_to_append = pd.DataFrame()
                    item_to_append = items_df
                    items_df = pd.concat([items_df, item_to_append], axis=0)
                elif (len(items_df.index) > (df_highest[0] / 2)):
                    item_to_append = pd.DataFrame()
                    df_to_add = df_highest[0] - len(items_df.index)
                    item_to_append = items_df[-df_to_add:]
                    items_df = pd.concat([items_df, item_to_append], axis=0)
                elif (len(items_df.index) == (df_highest[0])):
                    item_to_append = pd.DataFrame()
                    df_to_add = df_highest[0] - len(items_df.index)
                    item_to_append = items_df[-df_to_add:]
                    print(len(item_to_append.index))
                    items_df = pd.concat([items_df, item_to_append], axis=0)
                else:
                    pass
            df_processed = pd.concat([df_processed, items_df], axis=0)
        """Saving the balanced dataset"""
        path_to_save = os.getcwd() + '/Data Sets/' + column_name + '_Prediction/Train_' + column_name + '.xlsx'
        df_processed.to_excel(path_to_save, index=False)
        return df_processed.head()

    # method to map unique values to test dataset

    def create_test_file(self, input_data, column_name, column_name_modified, header):
        """
        :param input_data: input dataframe
        :param column_name: column to process
        :param column_name_modified: standard values
        :param header: Header of the column
        :return: first five values of the dataframe
        """
        self.input_data = input_data
        self.column_name = column_name
        self.column_name_modified = column_name_modified
        self.header = header

        """Creating empty test dataframe with headers"""
        test_column_name = pd.DataFrame(columns=[column_name, column_name_modified])
        """Adding values from input dataframe"""
        test_column_name[column_name] = input_data[header]
        """For cases where the column is not in string format, converting to string"""
        test_column_name[column_name] = test_column_name[column_name].astype(str)
        """performing basic preprocessing"""
        test_column_name.rename(columns=lambda x: x.strip())
        test_column_name = test_column_name.drop(test_column_name[(test_column_name[column_name] == '')].index)
        test_column_name = test_column_name.drop(test_column_name[(test_column_name[column_name] == ' ')].index)
        test_column_name.drop_duplicates(subset=[column_name], keep='first', inplace=True)
        test_column_name[column_name] = [word.lower() for word in test_column_name[column_name]]

        """Saving the test dataframe in a excel file"""
        path_to_save = os.getcwd() + '/Data Sets/' + column_name + '_Prediction/Test_' + column_name + '.xlsx'
        test_column_name.to_excel(path_to_save, index=False)
        return test_column_name.head()

    # method to read and clean train and test files.

    def read_train_file(self, column_name):
        """
        :param column_name: column to process
        :return: returning training dataframe
        """
        self.column_name = column_name
        """reading the train file"""
        path_to_open = os.getcwd() + '/Data Sets/' + column_name + '_Prediction/Train_' + column_name + '.xlsx'
        train_column_name = pd.read_excel(path_to_open, sheet_name='Sheet1', encoding='utf-8', na_filter=False,
                                          header=0)
        """Performing basic preprocessing"""
        train_column_name[column_name] = train_column_name[column_name].map(lambda x: str(x).strip())
        train_column_name[column_name] = [word.lower() for word in train_column_name[column_name]]
        return train_column_name

    def read_test_file(self, column_name):
        """
        :param column_name: column to process
        :return: return test dataframe
        """
        self.column_name = column_name
        """reading test file into a dataframe"""
        path_to_open = os.getcwd() + '/Data Sets/' + column_name + '_Prediction/Test_' + column_name + '.xlsx'
        test_column_name = pd.read_excel(path_to_open, sheet_name='Sheet1', encoding='utf-8', na_filter=False, header=0)

        """Performing basic preprocessing"""
        test_column_name[column_name] = test_column_name[column_name].apply(str)
        test_column_name.rename(columns=lambda x: x.strip())

        test_column_name[column_name] = test_column_name[column_name].map(lambda x: x.strip())
        test_column_name[column_name] = [word.lower() for word in test_column_name[column_name]]
        test_column_name = test_column_name.drop(test_column_name[(test_column_name[column_name] == '')].index)
        test_column_name = test_column_name.drop(test_column_name[(test_column_name[column_name] == ' ')].index)
        test_column_name.drop_duplicates(subset=[column_name], keep='first', inplace=True)
        return test_column_name

    # Vectorizing the text columns

    def vectorize_data(self, column_name, train_column_name, test_column_name):
        """
        :param column_name: column to process
        :param train_column_name: training dataframe
        :param test_column_name: test dataframe
        :return: returning tfidf transformed column values
        """
        self.column_name = column_name
        self.train_column_name = train_column_name
        self.test_column_name = test_column_name
        """implementing count vectorizer on training column"""
        count_vect = CountVectorizer()
        X_train_counts = count_vect.fit_transform(train_column_name[column_name])
        X_train_counts.shape
        """Implementing tfidf transfomer on vectorized column"""
        tfidf_transformer = TfidfTransformer()
        X_train_tfidf = tfidf_transformer.fit_transform(X_train_counts)
        return X_train_tfidf

    # NB model
    def validation_nb_model_fit(self, X_train, y_train, X_val, X_train_tfidf):
        """
        :param X_train: column to train on
        :param y_train: standard values to train on
        :param X_val: validation column
        :param X_train_tfidf: vectorized column
        :return: predictions
        """
        self.X_train = X_train
        self.y_train = y_train
        self.X_val = X_val
        self.X_train_tfidf = X_train_tfidf
        # Pipelining
        """processing the data over the pipeline wih naive bayes"""
        text_clf = Pipeline([('vect', CountVectorizer()), ('tfidf', TfidfTransformer()), ('clf', MultinomialNB()), ])
        text_clf = text_clf.fit(X_train, y_train)

        # using GridSearch CV
        parameters = {'vect__ngram_range': [(1, 1), (1, 2)], 'tfidf__use_idf': (True, False),
                      'clf__alpha': (1e-2, 1e-3), }
        gs_clf = GridSearchCV(text_clf, parameters, n_jobs=-1)
        gs_clf = gs_clf.fit(X_train, y_train)
        gs_clf.best_score_
        gs_clf.best_params_
        predicted = gs_clf.predict(X_val)
        return predicted

    def nb_model_fit(self, column_name, column_name_modified, train_column_name, test_column_name, X_train_tfidf):
        """
        :param column_name: column to process
        :param column_name_modified: standard values
        :param train_column_name: training dataframe
        :param test_column_name: test dataframe
        :param X_train_tfidf: vectorized values
        :return: predictions
        """
        self.column_name = column_name
        self.column_name_modified = column_name_modified
        self.train_column_name = train_column_name
        self.test_column_name = test_column_name
        self.X_train_tfidf = X_train_tfidf

        # Pipelining
        """processing the data over the pipeline wih naive bayes"""
        text_clf = Pipeline([('vect', CountVectorizer()), ('tfidf', TfidfTransformer()), ('clf', MultinomialNB()), ])
        text_clf = text_clf.fit(train_column_name[column_name], train_column_name[column_name_modified])
        """Fine tuning parameters"""
        parameters = {'vect__ngram_range': [(1, 1), (1, 2)], 'tfidf__use_idf': (True, False),
                      'clf__alpha': (1e-2, 1e-3), }
        gs_clf = GridSearchCV(text_clf, parameters, n_jobs=-1)
        gs_clf = gs_clf.fit(train_column_name[column_name], train_column_name[column_name_modified])
        gs_clf.best_score_
        gs_clf.best_params_
        predicted = gs_clf.predict(test_column_name[column_name])
        return predicted

    def nb_model_predict(self, column_name, column_name_modified, train_column_name, test_column_name, predicted,
                         all_column_dataframe):
        """
        :param column_name: column to be processed
        :param column_name_modified: standard values
        :param train_column_name: train datframe
        :param test_column_name: test dataframe
        :param predicted: prediction from the model
        :param all_column_dataframe: dataframe with all the columns together
        :return: mean score
        """
        self.column_name = column_name
        self.column_name_modified = column_name_modified
        self.train_column_name = train_column_name
        self.test_column_name = test_column_name
        self.predicted = predicted

        test_column_name['Predict_NB'] = predicted
        Save_Test = pd.DataFrame()
        all_column_dataframe[column_name] = test_column_name[column_name]
        all_column_dataframe['Predict_NB'] = test_column_name['Predict_NB']
        Save_Test[column_name] = test_column_name[column_name]
        Save_Test['Predict_NB'] = test_column_name['Predict_NB']
        path_to_save = os.getcwd() + '/Data Sets/' + column_name + '_Prediction/Predict_NB.xlsx'
        Save_Test.to_excel(path_to_save, index=False)
        return np.mean(Save_Test['Predict_NB'] == test_column_name[column_name])

    # Decision Tree model
    def validation_dt_model_fit(self, X_train, y_train, X_val, X_train_tfidf):
        """
        :param X_train:column to train on
        :param y_train:standard values to train on
        :param X_val:validation column
        :param X_train_tfidf:vectorized column
        :return:predictions
        """
        self.X_train = X_train
        self.y_train = y_train
        self.X_val = X_val
        self.X_train_tfidf = X_train_tfidf
        # Pipelining 
        dt = Pipeline([('vect', CountVectorizer()), ('tfidf', TfidfTransformer()),
                       ('clf-dt', DecisionTreeClassifier(criterion="gini", splitter="best",
                                                         max_depth=20, random_state=42)), ])
        _ = dt.fit(X_train, y_train)

        predicted = dt.predict(X_val)
        return predicted

    def dt_model_fit(self, column_name, column_name_modified, train_column_name, test_column_name, X_train_tfidf):
        """
        :param column_name:column to process
        :param column_name_modified:standard values
        :param train_column_name:training dataframe
        :param test_column_name:test dataframe
        :param X_train_tfidf:vectorized values
        :return:predictions
        """
        self.column_name = column_name
        self.column_name_modified = column_name_modified
        self.train_column_name = train_column_name
        self.test_column_name = test_column_name
        self.X_train_tfidf = X_train_tfidf
        dt = Pipeline([('vect', CountVectorizer()), ('tfidf', TfidfTransformer()),
                       ('clf-dt', DecisionTreeClassifier(criterion="gini", splitter="best",
                                                         max_depth=30, random_state=50)), ])
        _ = dt.fit(train_column_name[column_name], train_column_name[column_name_modified])
        predicted = dt.predict(test_column_name[column_name])
        return predicted

    def dt_model_predict(self, column_name, column_name_modified, train_column_name, test_column_name, predicted,
                         all_column_dataframe):
        """
        :param column_name:column to be processed
        :param column_name_modified:standard values
        :param train_column_name:train datframe
        :param test_column_name:test dataframe
        :param predicted:prediction from the model
        :param all_column_dataframe:dataframe with all the columns together
        :return:mean score
        """
        self.column_name = column_name
        self.column_name_modified = column_name_modified
        self.train_column_name = train_column_name
        self.test_column_name = test_column_name
        self.predicted = predicted
        test_column_name['Predict_DT'] = predicted
        Save_Test = pd.DataFrame()
        # Save_Test['Email_Address'] = test_column_name['Email_Address']
        Save_Test[column_name] = test_column_name[column_name]
        # Save_Test[column_name_modified] = test_column_name[column_name_modified]
        Save_Test['Predict_DT'] = test_column_name['Predict_DT']
        all_column_dataframe['Predict_DT'] = test_column_name['Predict_DT']
        path_to_save = os.getcwd() + '/Data Sets/' + column_name + '_Prediction/Predict_DT.xlsx'
        Save_Test.to_excel(path_to_save, index=False)
        return np.mean(Save_Test['Predict_DT'] == test_column_name[column_name])

    # Random Forest model
    def validation_rf_model_fit(self, X_train, y_train, X_val, X_train_tfidf):
        """
        :param X_train:column to train on
        :param y_train:standard values to train on
        :param X_val:validation column
        :param X_train_tfidf:vectorized column
        :return:predictions
        """
        self.X_train = X_train
        self.y_train = y_train
        self.X_val = X_val
        self.X_train_tfidf = X_train_tfidf
        # Pipelining
        rf = Pipeline([('vect', CountVectorizer()), ('tfidf', TfidfTransformer()),
                       ('clf-rf', RandomForestClassifier(n_estimators=100, max_depth=5, random_state=42)), ])
        _ = rf.fit(X_train, y_train)
        predicted = rf.predict(X_val)
        return predicted

    def rf_model_fit(self, column_name, column_name_modified, train_column_name, test_column_name, X_train_tfidf):
        """
        :param column_name:column to process
        :param column_name_modified:standard values
        :param train_column_name:training dataframe
        :param test_column_name:test dataframe
        :param X_train_tfidf:vectorized values
        :return:predictions
        """
        self.column_name = column_name
        self.column_name_modified = column_name_modified
        self.train_column_name = train_column_name
        self.test_column_name = test_column_name
        self.X_train_tfidf = X_train_tfidf
        rf = Pipeline([('vect', CountVectorizer()), ('tfidf', TfidfTransformer()),
                       ('clf-rf', RandomForestClassifier(n_estimators=100, max_depth=5, random_state=42)), ])
        _ = rf.fit(train_column_name[column_name], train_column_name[column_name_modified])
        predicted = rf.predict(test_column_name[column_name])
        return predicted

    def rf_model_predict(self, column_name, column_name_modified, train_column_name, test_column_name, predicted,
                         all_column_dataframe):
        """
        :param column_name:column to be processed
        :param column_name_modified:standard values
        :param train_column_name:train datframe
        :param test_column_name:test dataframe
        :param predicted:prediction from the model
        :param all_column_dataframe:dataframe with all the columns together
        :return:mean score
        """
        self.column_name = column_name
        self.column_name_modified = column_name_modified
        self.train_column_name = train_column_name
        self.test_column_name = test_column_name
        self.predicted = predicted
        test_column_name['Predict_RF'] = predicted
        Save_Test = pd.DataFrame()
        # Save_Test['Email_Address'] = test_column_name['Email_Address']
        Save_Test[column_name] = test_column_name[column_name]
        # Save_Test[column_name_modified] = test_column_name[column_name_modified]
        Save_Test['Predict_RF'] = test_column_name['Predict_RF']
        all_column_dataframe['Predict_RF'] = test_column_name['Predict_RF']
        path_to_save = os.getcwd() + '/Data Sets/' + column_name + '_Prediction/Predict_RF.xlsx'
        Save_Test.to_excel(path_to_save, index=False)
        return np.mean(Save_Test['Predict_RF'] == test_column_name[column_name])

    # SVM model
    def validation_svm_model_fit(self, X_train, y_train, X_val, X_train_tfidf):
        """
        :param X_train:column to train on
        :param y_train:standard values to train on
        :param X_val:validation column
        :param X_train_tfidf:vectorized column
        :return:predictions
        """
        self.X_train = X_train
        self.y_train = y_train
        self.X_val = X_val
        self.X_train_tfidf = X_train_tfidf
        # using SVM
        text_clf_svm = Pipeline([('vect', CountVectorizer()), ('tfidf', TfidfTransformer()),
                                 ('clf-svm', SGDClassifier(loss='hinge', penalty='l2', alpha=1e-3, max_iter=5,
                                                           random_state=42)), ])
        _ = text_clf_svm.fit(X_train, y_train)
        predicted_svm = text_clf_svm.predict(X_val)
        return predicted_svm

    def svm_model_fit(self, column_name, column_name_modified, train_column_name, test_column_name, X_train_tfidf):
        """
        :param column_name:column to process
        :param column_name_modified:standard values
        :param train_column_name:training dataframe
        :param test_column_name:test dataframe
        :param X_train_tfidf:vectorized values
        :return:predictions
        """
        self.column_name = column_name
        self.column_name_modified = column_name_modified
        self.train_column_name = train_column_name
        self.test_column_name = test_column_name
        self.X_train_tfidf = X_train_tfidf
        # using SVM
        text_clf_svm = Pipeline([('vect', CountVectorizer()), ('tfidf', TfidfTransformer()),
                                 ('clf-svm', SGDClassifier(loss='hinge', penalty='l2', alpha=1e-3, max_iter=5,
                                                           random_state=42)), ])
        _ = text_clf_svm.fit(train_column_name[column_name], train_column_name[column_name_modified])
        predicted_svm = text_clf_svm.predict(test_column_name[column_name])
        return predicted_svm

    def svm_model_predict(self, column_name, column_name_modified, train_column_name, test_column_name, predicted_svm,
                          all_column_dataframe):
        """
        :param column_name:column to be processed
        :param column_name_modified:standard values
        :param train_column_name:train datframe
        :param test_column_name:test dataframe
        :param predicted_svm:prediction from the model
        :param all_column_dataframe:dataframe with all the columns together
        :return:mean score
        """
        self.column_name = column_name
        self.column_name_modified = column_name_modified
        self.train_column_name = train_column_name
        self.test_column_name = test_column_name
        self.predicted_svm = predicted_svm
        test_column_name['Predict_SVM'] = predicted_svm
        Save_Test = pd.DataFrame()
        # Save_Test['Email_Address'] = test_column_name['Email_Address']
        Save_Test[column_name] = test_column_name[column_name]
        # Save_Test[column_name_modified] = test_column_name[column_name_modified]
        Save_Test['Predict_SVM'] = test_column_name['Predict_SVM']
        all_column_dataframe['Predict_SVM'] = test_column_name['Predict_SVM']
        path_to_save = os.getcwd() + '/Data Sets/' + column_name + '_Prediction/Predict_SVM.xlsx'
        Save_Test.to_excel(path_to_save, index=False)
        return np.mean(Save_Test['Predict_SVM'] == test_column_name[column_name])

    # SVM GS model
    def validation_svm_grid_search_model_fit(self, X_train, y_train, X_val, X_train_tfidf):
        """
        :param X_train:column to train on
        :param y_train:standard values to train on
        :param X_val:validation column
        :param X_train_tfidf:vectorized column
        :return:predictions
        """
        self.X_train = X_train
        self.y_train = y_train
        self.X_val = X_val
        self.X_train_tfidf = X_train_tfidf
        text_clf_svm = Pipeline([('vect', CountVectorizer()), ('tfidf', TfidfTransformer()),
                                 ('clf-svm', SGDClassifier(loss='hinge', penalty='l2', alpha=1e-3, max_iter=5,
                                                           random_state=42)), ])
        _ = text_clf_svm.fit(X_train, y_train)
        # using Grid Search CV with SVM
        parameters_svm = {'vect__ngram_range': [(1, 1), (1, 2)], 'tfidf__use_idf': (True, False),
                          'clf-svm__alpha': (1e-2, 1e-3), }
        gs_clf_svm = GridSearchCV(text_clf_svm, parameters_svm, n_jobs=-1)
        gs_clf_svm = gs_clf_svm.fit(X_train, y_train)
        predicted_gs_clf_svm = gs_clf_svm.predict(X_val)
        gs_clf_svm.best_score_
        return predicted_gs_clf_svm

    def svm_grid_search_model_fit(self, column_name, column_name_modified,
                                  train_column_name, test_column_name, X_train_tfidf):
        """
        :param column_name:column to process
        :param column_name_modified:standard values
        :param train_column_name:training dataframe
        :param test_column_name:test dataframe
        :param X_train_tfidf:vectorized values
        :return:predictions
        """
        self.column_name = column_name
        self.column_name_modified = column_name_modified
        self.train_column_name = train_column_name
        self.test_column_name = test_column_name
        self.X_train_tfidf = X_train_tfidf
        text_clf_svm = Pipeline([('vect', CountVectorizer()), ('tfidf', TfidfTransformer()),
                                 ('clf-svm', SGDClassifier(loss='hinge', penalty='l2', alpha=1e-3, max_iter=5,
                                                           random_state=42)), ])
        _ = text_clf_svm.fit(train_column_name[column_name], train_column_name[column_name_modified])
        # using Grid Search CV with SVM
        parameters_svm = {'vect__ngram_range': [(1, 1), (1, 2)], 'tfidf__use_idf': (True, False),
                          'clf-svm__alpha': (1e-2, 1e-3), }
        gs_clf_svm = GridSearchCV(text_clf_svm, parameters_svm, n_jobs=-1)
        gs_clf_svm = gs_clf_svm.fit(train_column_name[column_name], train_column_name[column_name_modified])
        predicted_gs_clf_svm = gs_clf_svm.predict(test_column_name[column_name])
        gs_clf_svm.best_score_
        return predicted_gs_clf_svm

    def svm_grid_search_model_predict(self, column_name, column_name_modified, train_column_name, test_column_name,
                                      predicted_gs_clf_svm, all_column_dataframe):
        """
        :param column_name:column to be processed
        :param column_name_modified:standard values
        :param train_column_name:train datframe
        :param test_column_name:test dataframe
        :param predicted_gs_clf_svm:prediction from the model
        :param all_column_dataframe:dataframe with all the columns together
        :return:mean score
        """
        self.column_name = column_name
        self.column_name_modified = column_name_modified
        self.train_column_name = train_column_name
        self.test_column_name = test_column_name
        self.predicted_gs_clf_svm = predicted_gs_clf_svm
        test_column_name['Predict_SVM_GS'] = predicted_gs_clf_svm
        Save_Test = pd.DataFrame()
        # Save_Test['Email_Address'] = test_column_name['Email_Address']
        Save_Test[column_name] = test_column_name[column_name]
        # Save_Test[column_name_modified] = test_column_name[column_name_modified]
        Save_Test['Predict_SVM_GS'] = test_column_name['Predict_SVM_GS']
        all_column_dataframe['Predict_SVM_GS'] = test_column_name['Predict_SVM_GS']
        path_to_save = os.getcwd() + '/Data Sets/' + column_name + '_Prediction/Predict_SVM_GS.xlsx'
        Save_Test.to_excel(path_to_save, index=False)
        return np.mean(Save_Test['Predict_SVM_GS'] == test_column_name[column_name])

    # nb nltk model
    def validation_nltk_nb_model_fit(self, X_train, y_train, X_val, X_train_tfidf):
        """
        :param X_train:column to train on
        :param y_train:standard values to train on
        :param X_val:validation column
        :param X_train_tfidf:vectorized column
        :return:predictions
        """
        self.X_train = X_train
        self.y_train = y_train
        self.X_val = X_val
        self.X_train_tfidf = X_train_tfidf
        # nltk.download()
        stemmer = SnowballStemmer("english", ignore_stopwords=True)

        class StemmedCountVectorizer(CountVectorizer):
            def build_analyzer(self):
                analyzer = super(StemmedCountVectorizer, self).build_analyzer()
                return lambda doc: ([stemmer.stem(w) for w in analyzer(doc)])

        stemmed_count_vect = StemmedCountVectorizer(stop_words='english')
        text_mnb_stemmed = Pipeline(
            [('vect', stemmed_count_vect), ('tfidf', TfidfTransformer()), ('mnb', MultinomialNB(fit_prior=False)), ])
        text_mnb_stemmed = text_mnb_stemmed.fit(X_train, y_train)
        predicted_mnb_stemmed = text_mnb_stemmed.predict(X_val)
        return predicted_mnb_stemmed

    def nltk_nb_model_fit(self, column_name, column_name_modified, train_column_name, test_column_name, X_train_tfidf):
        """
        :param column_name:column to process
        :param column_name_modified:standard values
        :param train_column_name:training dataframe
        :param test_column_name:test dataframe
        :param X_train_tfidf:vectorized values
        :return:predictions
        """
        self.column_name = column_name
        self.column_name_modified = column_name_modified
        self.train_column_name = train_column_name
        self.test_column_name = test_column_name
        self.X_train_tfidf = X_train_tfidf
        # nltk.download()
        stemmer = SnowballStemmer("english", ignore_stopwords=True)

        class StemmedCountVectorizer(CountVectorizer):
            def build_analyzer(self):
                analyzer = super(StemmedCountVectorizer, self).build_analyzer()
                return lambda doc: ([stemmer.stem(w) for w in analyzer(doc)])

        stemmed_count_vect = StemmedCountVectorizer(stop_words='english')
        text_mnb_stemmed = Pipeline(
            [('vect', stemmed_count_vect), ('tfidf', TfidfTransformer()), ('mnb', MultinomialNB(fit_prior=False)), ])
        text_mnb_stemmed = text_mnb_stemmed.fit(train_column_name[column_name], train_column_name[column_name_modified])
        predicted_mnb_stemmed = text_mnb_stemmed.predict(test_column_name[column_name])
        return predicted_mnb_stemmed

    def nltk_nb_model_predict(self, column_name, column_name_modified, train_column_name, test_column_name,
                              predicted_mnb_stemmed, all_column_dataframe):
        """
        :param column_name:column to be processed
        :param column_name_modified:standard values
        :param train_column_name:train datframe
        :param test_column_name:test dataframe
        :param predicted_mnb_stemmed:prediction from the model
        :param all_column_dataframe:dataframe with all the columns together
        :return:mean score
        """
        self.column_name = column_name
        self.column_name_modified = column_name_modified
        self.train_column_name = train_column_name
        self.test_column_name = test_column_name
        self.predicted_mnb_stemmed = predicted_mnb_stemmed
        test_column_name['Predict_NLTK_NB'] = predicted_mnb_stemmed
        Save_Test = pd.DataFrame()
        # Save_Test['Email_Address'] = test_column_name['Email_Address']
        Save_Test[column_name] = test_column_name[column_name]
        # Save_Test[column_name_modified] = test_column_name[column_name_modified]
        Save_Test['Predict_NLTK_NB'] = test_column_name['Predict_NLTK_NB']
        all_column_dataframe['Predict_NLTK_NB'] = test_column_name['Predict_NLTK_NB']
        path_to_save = os.getcwd() + '/Data Sets/' + column_name + '_Prediction/Predict_NLTK_NB.xlsx'
        Save_Test.to_excel(path_to_save, index=False)

        return np.mean(Save_Test['Predict_NLTK_NB'] == test_column_name[column_name])

    # generating classification report for the above data
    def classification_report_nb(self, y_val, predicted):  # target_name_for_classification,
        """
        :param y_val: actual value from test dataset
        :param predicted: predicted value from test dataset
        :return: accuracy score
        """
        self.y_val = y_val
        self.predicted = predicted
        precision_nb = accuracy_score(y_val, predicted)
        return precision_nb

    def classification_report_svm(self, y_val, predicted_svm):  # target_name_for_classification,
        """
        :param y_val: actual value from test dataset
        :param predicted_svm: predicted value from test dataset
        :return:accuracy score
        """
        self.y_val = y_val
        self.predicted_svm = predicted_svm
        precision_svm = accuracy_score(y_val, predicted_svm)
        return precision_svm

    def classification_report_gs_svm(self, y_val, predicted_gs_clf_svm):  # target_name_for_classification,
        """
        :param y_val: actual value from test dataset
        :param predicted_gs_clf_svm:predicted value from test dataset
        :return:accuracy score
        """
        self.y_val = y_val
        self.predicted_gs_clf_svm = predicted_gs_clf_svm
        precision_gs_svm = accuracy_score(y_val, predicted_gs_clf_svm)
        return precision_gs_svm

    def classification_report_nltk_nb(self, y_val, predicted_mnb_stemmed):  # target_name_for_classification,
        """
        :param y_val: actual value from test dataset
        :param predicted_mnb_stemmed:predicted value from test dataset
        :return:accuracy score
        """
        self.y_val = y_val
        self.predicted_mnb_stemmed = predicted_mnb_stemmed
        precision_nltk_nb = accuracy_score(y_val, predicted_mnb_stemmed)
        return precision_nltk_nb

    def classification_report_rf(self, y_val, predicted_rf):  # target_name_for_classification,
        """
        :param y_val: actual value from test dataset
        :param predicted_rf: predicted value from test dataset
        :return: accuracy score
        """
        self.y_val = y_val
        self.predicted_mnb_stemmed = predicted_rf
        precision_rf = accuracy_score(y_val, predicted_rf)
        return precision_rf

    def classification_report_dt(self, y_val, predicted_dt):  # target_name_for_classification,
        """
        :param y_val: actual value from test dataset
        :param predicted_dt: predicted value from test dataset
        :return: accuracy score
        """
        self.y_val = y_val
        self.predicted_mnb_stemmed = predicted_dt
        precision_dt = accuracy_score(y_val, predicted_dt)
        return precision_dt

    # creating confusion matrix to check the quality of model

    def confusion_matrix_nb(self, column_name, column_name_modified, train_column_name,
                            test_column_name, target_name_for_classification, predicted):
        """
        :param column_name: column trained on
        :param column_name_modified:standard values to train on
        :param train_column_name:column name trained
        :param test_column_name:column name for test
        :param target_name_for_classification:labels used for classification
        :param predicted:predicted values
        :return:confusion matrix with positives and negatives
        """
        self.column_name = column_name
        self.column_name_modified = column_name_modified
        self.train_column_name = train_column_name
        self.test_column_name = test_column_name
        self.target_name_for_classification = target_name_for_classification
        self.predicted = predicted
        return confusion_matrix(test_column_name[column_name_modified], predicted,
                                labels=target_name_for_classification)

    def confusion_matrix_svm(self, column_name, column_name_modified, train_column_name,
                             test_column_name, target_name_for_classification, predicted_svm):
        """
        :param column_name: column trained on
        :param column_name_modified: standard values to train on
        :param train_column_name: column name trained
        :param test_column_name: column name for test
        :param target_name_for_classification:labels used for classification
        :param predicted_svm: predicted values
        :return:confusion matrix with positives and negatives
        """
        self.column_name = column_name
        self.column_name_modified = column_name_modified
        self.train_column_name = train_column_name
        self.test_column_name = test_column_name
        self.target_name_for_classification = target_name_for_classification
        self.predicted_svm = predicted_svm
        return confusion_matrix(test_column_name[column_name_modified], predicted_svm,
                                labels=target_name_for_classification)

    def confusion_matrix_gs_svm(self, column_name, column_name_modified, train_column_name,
                                test_column_name, target_name_for_classification, predicted_gs_clf_svm):
        """
        :param column_name: column trained on
        :param column_name_modified: standard values to train on
        :param train_column_name: column name trained
        :param test_column_name: column name for test
        :param target_name_for_classification: labels used for classification
        :param predicted_gs_clf_svm: predicted values
        :return: confusion matrix with positives and negatives
        """
        self.column_name = column_name
        self.column_name_modified = column_name_modified
        self.train_column_name = train_column_name
        self.test_column_name = test_column_name
        self.target_name_for_classification = target_name_for_classification
        self.predicted_gs_clf_svm = predicted_gs_clf_svm
        return confusion_matrix(test_column_name[column_name_modified], predicted_gs_clf_svm,
                                labels=target_name_for_classification)

    def confusion_matrix_nltk_nb(self, column_name, column_name_modified, train_column_name,
                                 test_column_name, target_name_for_classification, predicted_mnb_stemmed):
        """
        :param column_name: column trained on
        :param column_name_modified: standard values to train on
        :param train_column_name: column name trained
        :param test_column_name: column name for test
        :param target_name_for_classification: labels used for classification
        :param predicted_mnb_stemmed: predicted values
        :return: confusion matrix with positives and negatives
        """
        self.column_name = column_name
        self.column_name_modified = column_name_modified
        self.train_column_name = train_column_name
        self.test_column_name = test_column_name
        self.target_name_for_classification = target_name_for_classification
        self.predicted_mnb_stemmed = predicted_mnb_stemmed
        return confusion_matrix(test_column_name[column_name_modified], predicted_mnb_stemmed,
                                labels=target_name_for_classification)

        # scoring each row based on its prediction

    def prediction_rate(self, column_name, column_name_modified, train_column_name,
                        test_column_name, model_name, file_name):
        """
        :param column_name: column trained on
        :param column_name_modified: standard values to train on
        :param train_column_name: column name trained
        :param test_column_name: column name for test
        :param model_name: name of the model selected for prediction
        :param file_name: file name to open the predicted values
        :return:
        """
        self.column_name = column_name
        self.column_name_modified = column_name_modified
        self.train_column_name = train_column_name
        self.test_column_name = test_column_name
        self.model_name = model_name
        self.file_name = file_name
        predict_model_name = 'Predict_' + model_name
        file_path = os.getcwd() + '/Data Sets/' + column_name + '_Prediction/' + file_name
        column_model_name = pd.read_excel(file_path, sheet_name='Sheet1', encoding='utf-8', na_filter=False,
                                          header=0)
        path_to_open_classification_report = os.getcwd() + '/Data Sets/' + column_name + '_Prediction/' + column_name + '_classification_report_' + column_name + '_' + model_name + '.xlsx'
        Selected_Value = 'Selected_Prediction'
        column_model_name.insert(loc=1, column=Selected_Value, value=column_model_name[predict_model_name])
        column_model_name.insert(loc=2, column='Prediction_Percent', value=-1)
        if os.path.isfile(path_to_open_classification_report):
            classification_report_scores = pd.read_excel(path_to_open_classification_report, sheet_name='Sheet1',
                                                         encoding='utf-8', na_filter=False, header=0)
            column_model_name.insert(loc=3, column='F1_Score', value=-0.01)
        column_model_name = column_model_name.drop([predict_model_name], axis=1)
        for index, words_in_a_row in column_model_name.iterrows():
            if os.path.isfile(path_to_open_classification_report):
                for f1_score_row in classification_report_scores.itertuples():
                    if words_in_a_row.Selected_Prediction.strip() == f1_score_row.Labels.strip():
                        column_model_name.loc[index, 'F1_Score'] = f1_score_row[2]
                        #column_model_name.set_value(words_in_a_row.Index, 'F1_Score', f1_score_row[2])
        return column_model_name

"""this class contains a method that is used for creating a classification report and saving it in an excel sheet"""
class Classification_Reports:
    def classifcation_report_processing(self, column_name, model_name, model_to_report):
        """
        :param column_name: column to process
        :param model_name: model name to create classification report
        :param model_to_report: classification report of the selected model
        :return: message if it is saved or not
        """
        self.column_name = column_name
        self.model_to_report = model_to_report
        self.model_name = model_name
        message = 'Classification Report Saved'
        tmp = list()
        """splitting the report on new line character"""
        for row in model_to_report.split("\n"):
            parsed_row = [x for x in row.split("  ") if len(x) > 0]
            if len(parsed_row) > 0:
                tmp.append(parsed_row)
        # Store in dictionary
        measures = tmp[0]
        try:
            """saving the data in a dictionary and then finally to a dataframe with transpose"""
            D_class_data = defaultdict(dict)
            for row in tmp[1:]:
                class_label = row[0]
                for j, m in enumerate(measures):
                    D_class_data[class_label][m.strip()] = float(row[j + 1].strip())
            save_report = pd.DataFrame.from_dict(D_class_data).T
            save_report.index.name = 'Labels'
            path_to_save = os.getcwd() + '/Data Sets/' + column_name + '_Prediction/' + column_name + '_classification_report_' + column_name + '_' + model_name + '.xlsx'
            """Finally saving the classification report"""
            save_report.to_excel(path_to_save, index=True)
        except IndexError:
            message = 'Classification cannot be created for' + column_name + ' ' + model_name
        return message

# # Class to run the process of text processing by taking all the inputs and carry out human interaction with machine
# The method to run all the functions of machine learning and non-machine learning process
class Process_All_Text_Columns:
    def ml_process_general_text_columns_processing(self, input_file_name, input_file_sheet_name):
        self.input_file_name = input_file_name
        self.input_file_sheet_name = input_file_sheet_name
        column_prediction_sorting_order_input = '1'
        mapped_file = os.getcwd() + '\\Data Sets\\Mapping_Technical_Data.xlsx'
        """
        :return: na
        """
        """get the current data time"""
        now = datetime.datetime.now()
        """setting the name of log file with current date time
        then setting the basic level of logging"""
        #path_for_logs = os.getcwd() + '\\Log_Folder'
        #if (os.path.isdir(path_for_logs)):
        #    pass
        #else:
        #    os.mkdir(path_for_logs)
        #log_file_name = path_for_logs + '\\ML Application log file created on ' + now.strftime("%Y-%m-%d "
                                                                                                     # "%H-%M") + '.log'
        #logging.basicConfig(filename=log_file_name, level=logging.INFO)
        #logging.info(r'Machine Learning Application started with User Interface.')
        #logging.info('!!!Important Information!!! All files will be saved at ', os.getcwd(),
                     #'\\Data Sets\\', ' , Please check the dirctory for the files created.')

        #logging.info('User Entered mandatory Fields, next steps will check if the entered files exists.')
        """getting the names of all the columns in a list"""
        column_name_list = ['Initial_Value']
        print(column_name_list)
        #mapping_file_sheet_name = 'mappingsheet'
        """running over all the columns user has entered for machine learning process"""
        for columns_array in column_name_list:
            """If the training file already exists then train the training file using the predicted vlaues"""
            import_train_file_mapping = Unique_Values_To_Training_File()
            import_train_file_mapping.Train_The_Training_File(columns_array, input_file_name)
            input_file_path = '0'
            while (input_file_path == "0"):
                #logging.info(input_file_name)
                print(input_file_name)
                input_data = pd.DataFrame()
                """checking if the the file path entered by the user exists or not"""
                if (os.path.isfile(input_file_name)):
                    input_file_name_exists = '0'
                    while (input_file_name_exists == '0'):
                        """checking for file if it is csv or xlsx file to read else error message printed as format 
                        not supported"""
                        if input_file_name.rpartition('.')[2] == 'pdf':
                            with open(input_file_name, 'rb') as f:
                                pages = PdfFileReader(f).getNumPages()

                            tables = camelot.read_pdf(input_file_name, pages='all')
                            input_data = pd.DataFrame
                            dataframe_page = pd.DataFrame

                            for page in range(pages):
                                if page == 0:
                                    if pages == 1:
                                        tables = camelot.read_pdf(input_file_name)
                                        input_data = tables[0].df
                                        break
                                    else:
                                        input_data = tables[page].df
                                else:
                                    dataframe_page = tables[page].df
                                    input_data.rename(columns={list(input_data)[0]: 'Initial_Value'}, inplace=True)
                                    dataframe_page.rename(columns={list(dataframe_page)[0]: 'Initial_Value'},
                                                          inplace=True)

                                    if dataframe_page['Initial_Value'].equals(input_data['Initial_Value']):
                                        dataframe_page = dataframe_page.drop(columns='Initial_Value', axis=1)
                                        dataframe_page = dataframe_page.drop(columns=1, axis=1)
                                        input_data = pd.concat([input_data, dataframe_page], axis=1)
                                    else:
                                        input_data = pd.concat([input_data, dataframe_page], axis=0)
                                print(input_data)
                                input_file_name_exists = '1'
                                if pages == page:
                                    break
                            print('Saving PDF as excel for further cleaning.')
                            path_to_save = os.getcwd() + '/Data Sets/Pre_Cleaned_File.xlsx'
                            input_data.to_excel(path_to_save, index=False)
                            doClean = DoCleanInput()
                            input_data = doClean.do_cleaning(mapped_file, path_to_save, 'Sheet1')
                        else :
                            #doTrans = DoTranspose()
                            #input_data = doTrans.do_transposer(mapped_file, input_file_name, input_file_sheet_name)
                            # if flag_trans == 1:
                            #     print('Input file needed transpose, transposed and saved, processing further.')
                            #     path_to_open = os.getcwd() + '/Data Sets/Transposed_File.xlsx'
                            #     input_data = pd.read_xlsx(path_to_open, encoding='utf-8', header=0, na_filter=False)

                            if input_data.empty:
                                print('Input file dont need to be transformed, processing further.')
                                doClean = DoCleanInput()
                                input_data = doClean.do_cleaning(mapped_file, input_file_name, input_file_sheet_name)

                                # input_data = pd.read_csv(input_file_name, encoding='utf-8', header=0, na_filter=False,
                                #                          error_bad_lines=False)
                                # input_data.rename(columns={list(input_data)[0]: 'Initial_Value'}, inplace=True)
                            #logging.info('Loading Input file')
                            print(colored('Loading Input file', 'green'))
                            input_file_name_exists = '1'

                            #input_file_name_exists_message = args.input_file_name_exists_message
                            #if (input_file_name_exists_message == '0'):
                              #  input_file_path = '0'
                            #else:
                            #    break
                    #profile = pandas_profiling.ProfileReport(input_data)
                    #profile.to_file(os.getcwd() + '\\Data Sets\\' + "Input_File_Profile_Report.html")
                    print('Column Coming.. ', columns_array)
                    #logging.info('matching Headers with user selected column.')
                    header = 'Initial_Value'
                    input_file_path = '1'
                    column_name = columns_array
                    print()
                    """getting the list of header from input file and iterating over them"""
                #for header in header_names:
                    header_with_underscore = str(header).replace(" ", "_")
                    """if column name given by user exists in input file then do further processing"""
                    if (header == column_name or header_with_underscore == column_name):
                        #logging.info('Column Found')
                        directory_to_save_files = header
                        directory_to_save_files = directory_to_save_files.replace(" ", "_")
                        column_name = column_name.replace(" ", "_")
                        """Check if the directory exists else create new to save the files"""
                        data_set_path = os.getcwd() + '\\Data Sets\\'
                        if (os.path.isdir(data_set_path)):
                            #logging.info('Data Sets Directory Exists.')
                            pass
                        else:
                            #logging.info('Directory Created Data Sets.')
                            os.makedirs(data_set_path)
                        path = os.getcwd() + '\\Data Sets\\' + directory_to_save_files + "_Prediction"
                        # check if the directory for each of the file exists
                        """Check for each new column directory else create it"""
                        if (os.path.isdir(path)):
                            #logging.info('Column Folder Found')
                            pass
                        else:
                            #logging.info('Column Folder Created')
                            os.makedirs(path)
                        print(colored("Directory created, checking for Train Data set...!!!", 'green'))
                        import_text_processing_class = General_Methods_Class_for_Text_Processing()
                        column_name_modified = column_name + '_Modified'
                        """Calling the function to select the unique values from input file"""
                        column_name_head = import_text_processing_class.create_unique_values(input_data,
                                                                                             column_name,
                                                                                             column_name_modified,
                                                                                             header)
                        path_train_file = os.getcwd() + '\\Data Sets\\' + column_name + '_Prediction\\Train_' + column_name + '.xlsx'
                        path_test_file = os.getcwd() + '\\Data Sets\\' + column_name + '_Prediction\\Test_' + column_name + '.xlsx'
                        print(colored("Train Data set file created...!!!", 'green'))
                        #logging.info('Train Data set file created')

                        picklist_column_file_path = '0'

                        mapped_file_sheet_name = column_name
                        """checking if the label file exists as the input given by user"""
                        if (os.path.isfile(mapped_file)):
                            check_xl_sheet = pd.ExcelFile(mapped_file)
                            """reading it in either csv or xlsx format"""
                            if (mapped_file.rpartition('.')[2] == 'csv'):
                                train_column = pd.read_csv(mapped_file, encoding='utf-8',
                                                           header=0, na_filter=False)
                                #logging.info('Train File Created')
                                train_column.columns = [column_name, column_name_modified]
                                if not os.path.exists(os.getcwd() + '/Training_Files'):
                                    os.makedirs(os.getcwd() + '/Training_Files')
                                path_to_save = os.getcwd() + '/Data Sets/Training_Files/To_Process_Train_' + column_name + '.xlsx'
                                train_column.to_excel(path_to_save, index=False)
                            elif (mapped_file.rpartition('.')[2] == 'xlsx'):
                                train_column = pd.read_excel(mapped_file,
                                                             encoding='utf-8', na_filter=False, header=0)
                                #logging.info('Train File Created')
                                train_column = train_column.drop(["Unit", "Example"], axis=1)
                                train_column.columns = [column_name, column_name_modified]
                                if not os.path.exists(os.getcwd() + '/Data Sets/Training_Files'):
                                    os.makedirs(os.getcwd() + '/Data Sets/Training_Files')
                                path_to_save = os.getcwd() + '/Data Sets/Training_Files/To_Process_Train_' + column_name + '.xlsx'
                                train_column.to_excel(path_to_save, index=False)
                            elif mapped_file_sheet_name in check_xl_sheet.sheet_names:
                                train_column = pd.read_excel(mapped_file, sheet_name=mapped_file_sheet_name,
                                                             encoding='utf-8', na_filter=False, header=0)
                                #logging.info('Train File Created')
                                train_column = train_column.drop(["Unit", "Example"], axis=1)
                                train_column.columns = [column_name, column_name_modified]
                                if not os.path.exists(os.getcwd() + '/Data Sets/Training_Files'):
                                    os.makedirs(os.getcwd() + '/Data Sets/Training_Files')
                                path_to_save = os.getcwd() + '/Data Sets/Training_Files/To_Process_Train_' + column_name + '.xlsx'
                                train_column.to_excel(path_to_save, index=False)
                            else:
                                """else use the piclist file if label file doesnot exist to create the first 
                                train file"""
                                while (picklist_column_file_path == '0'):
                                    picklist_column_file = os.getcwd() + '\\Data Sets\\Mapping_Technical_Data.xlsx'
                                    if (os.path.isfile(picklist_column_file)):
                                        picklist_column_file_path = '1'
                                        picklist_column_file_exists = '0'
                                        while (picklist_column_file_exists == '0'):
                                            if (picklist_column_file.rpartition('.')[2] == 'csv'):
                                                picklist_column = pd.read_csv(picklist_column_file,
                                                                              encoding='utf-8',
                                                                              header=0, na_filter=False)
                                                picklist_column_file_exists = '1'
                                                #logging.info('Train File Created')
                                            elif (picklist_column_file.rpartition('.')[2] == 'xlsx'):
                                                #logging.info('Train File Created')
                                                picklist_column = pd.read_excel(picklist_column_file,
                                                                                sheet_name='Tabelle1',
                                                                                encoding='utf-8', na_filter=False,
                                                                                header=0)
                                                picklist_column_file_exists = '1'
                                            else:
                                                print(colored(
                                                    "file format not supported, please enter the file in .xlsx or .csv format.",
                                                    'red'))
                                                picklist_name_file_exists_message = colored(
                                                    "Press 0 to try again entering file path.\nPress any key to exit.",
                                                    'blue')
                                                #logging.info('Train File not found')


                                    else:
                                        print(colored("The file at the give path does not exist. Line Number 1363",
                                                      'red'))
                                        picklist_column_file_path_message_colored = colored(
                                            "Do you wish to try again.\nPress 0 to try again or any key to exit\nYour Response: ",
                                            'blue')
                                        #logging.info('Train File not found')
                                        picklist_column_file_path_message = '0'
                                        if (picklist_column_file_path_message == '0'):
                                            picklist_column_file_path = '0'
                                        else:
                                            break
                                if (os.path.isfile(path_train_file)):
                                    print("Training File Already Exists using the same.")
                                    #logging.info('Train File Exists')
                                else:
                                    picklist_data = import_text_processing_class.map_train_file_with_picklist(
                                        column_name,
                                        column_name_modified,
                                        header,
                                        picklist_column)
                                    display = colored(
                                        '1. Map the values which are marked as FILLED ==> FALSE in the '
                                        + header + ' file\n2. File path for the file:\n' +
                                        path + '\n3. File to edit:\nTo_Process_Train_' + column_name
                                        + '.xlsx\n4. Press Enter to continue when mapping is done.',
                                        'blue')
                                    #logging.info('Mapping Train File')
                        else:
                            while (picklist_column_file_path == '0'):
                                """if user didnot enter a lable file then use picklist file"""
                                picklist_column_file = os.getcwd() + 'Data Sets\\Mapping_Technical_Data.xlsx'
                                if (os.path.isfile(picklist_column_file)):
                                    picklist_column_file_path = '1'
                                    picklist_column_file_exists = '0'
                                    while (picklist_column_file_exists == '0'):
                                        if (picklist_column_file.rpartition('.')[2] == 'csv'):
                                            picklist_column = pd.read_csv(picklist_column_file, encoding='utf-8',
                                                                          header=0, na_filter=False)
                                            picklist_column_file_exists = '1'
                                            #logging.info('Train File Created')
                                        elif (picklist_column_file.rpartition('.')[2] == 'xlsx'):
                                            #logging.info('Train File Created')
                                            picklist_column = pd.read_excel(picklist_column_file,
                                                                            sheet_name='Tabelle1',
                                                                            encoding='utf-8', na_filter=False,
                                                                            header=0)
                                            picklist_column_file_exists = '1'
                                        else:
                                            print(colored(
                                                "file format not supported, please enter the file in .xlsx or .csv format.",
                                                'red'))
                                            picklist_name_file_exists_message = colored(
                                                "Press 0 to try again entering file path.\nPress any key to exit.",
                                                'blue')
                                            #logging.info('Train File not found')
                                else:
                                    print(colored("The file at the give path does not exist. Line Number 1419",
                                                  'red'))
                                    picklist_column_file_path_message_colored = colored(
                                        "Do you wish to try again.\nPress 0 to try again or any key to exit\nYour Response: ",
                                        'blue')
                                    #logging.info('Train File not found')
                                    picklist_column_file_path_message = '0'
                                    if (picklist_column_file_path_message == '0'):
                                        picklist_column_file_path = '0'
                                    else:
                                        break

                            """check if training file already exists else create a new training file"""
                            if (os.path.isfile(path_train_file)):
                                print("Training File Already Exists using the same.")
                                #logging.info('Train File Exists')
                            else:
                                picklist_data = import_text_processing_class.map_train_file_with_picklist(
                                    column_name,
                                    column_name_modified,
                                    header,
                                    picklist_column)
                                display = colored(
                                    '1. Map the values which are marked as FILLED ==> FALSE in the ' + header + ' file\n2. File path for the file:\n' +
                                    path + '\n3. File to edit:\nTo_Process_Train_' + column_name + '.xlsx\n4. Press Enter to continue when mapping is done.',
                                    'blue')
                                #logging.info('Mapping Train File')
                        #logging.info('Loading and cross Checking the Train File')
                        print(colored('Loading and cross Checking the Train File.', 'green'))
                        """balancing the training file before processing further"""
                        train_data_ready = import_text_processing_class.dataset_balancing_in_train_file(column_name,
                                                                                                        column_name_modified,
                                                                                                        header)
                        #logging.info('Train File check complete')
                        if (os.path.isfile(path_train_file)):
                            print(colored("Creating Test file.\nPlease wait progress will be updated below...!!!",
                                          'green'))
                        else:
                            message_train = path_train_file + "Files still not found, quitting the current column, if more column exists, the process will continue."
                            print(colored(message_train, 'red'))
                            break
                        """create new test file with unique values from input file"""
                        # if (os.path.isfile(path_test_file)):
                        #     pass
                        # else:
                        test_values_head = import_text_processing_class.create_test_file(input_data,
                                                                                         column_name,
                                                                                         column_name_modified,
                                                                                         header)
                        print(colored("Necessary files located, Machine Learning steps will follow shortly...!!!",
                                      'green'))
                        #logging.info('Test File Created')
                        """Read Training file and create validation set"""
                        train_column_name = import_text_processing_class.read_train_file(column_name)
                        X_train, X_val, y_train, y_val = train_test_split(train_column_name[column_name],
                                                                          train_column_name[column_name_modified],
                                                                          test_size=0.3, random_state=1)
                        test_column_name = import_text_processing_class.read_test_file(column_name)
                        X_train_tfidf = import_text_processing_class.vectorize_data(column_name, train_column_name,
                                                                                    test_column_name)

                        print(colored("Running different models:", 'green'))
                        #logging.info('Running Models')
                        all_column_dataframe = pd.DataFrame()
                        # Models
                        """Run the process through all the models Naive Bayes, SVM, Decision Tree, Random Forest
                        and get the accuracy score for each so that the best model can be selected"""
                        predict_val_NB = import_text_processing_class.validation_nb_model_fit(X_train, y_train,
                                                                                              X_val, X_train_tfidf)
                        predict_NB = import_text_processing_class.nb_model_fit(column_name, column_name_modified,
                                                                               train_column_name, test_column_name,
                                                                               X_train_tfidf)
                        score_NB = import_text_processing_class.nb_model_predict(column_name, column_name_modified,
                                                                                 train_column_name,
                                                                                 test_column_name,
                                                                                 predict_NB, all_column_dataframe)
                        classification_report_nb = import_text_processing_class.classification_report_nb(y_val,
                                                                                                         predict_val_NB)
                        print("Naive Bayes ", classification_report_nb)
                        #logging.info("Naive Bayes ", classification_report_nb)
                        predict_val_RF = import_text_processing_class.validation_rf_model_fit(X_train, y_train,
                                                                                              X_val, X_train_tfidf)
                        predict_RF = import_text_processing_class.rf_model_fit(column_name, column_name_modified,
                                                                               train_column_name, test_column_name,
                                                                               X_train_tfidf)
                        score_RF = import_text_processing_class.rf_model_predict(column_name, column_name_modified,
                                                                                 train_column_name,
                                                                                 test_column_name,
                                                                                 predict_RF, all_column_dataframe)
                        classification_report_rf = import_text_processing_class.classification_report_rf(y_val,
                                                                                                         predict_val_RF)
                        print("Random Forest ", classification_report_rf)
                        #logging.info("Random Forest ", classification_report_rf)
                        predict_val_DT = import_text_processing_class.validation_dt_model_fit(X_train, y_train,
                                                                                              X_val, X_train_tfidf)
                        predict_DT = import_text_processing_class.dt_model_fit(column_name, column_name_modified,
                                                                               train_column_name, test_column_name,
                                                                               X_train_tfidf)
                        score_DT = import_text_processing_class.dt_model_predict(column_name, column_name_modified,
                                                                                 train_column_name,
                                                                                 test_column_name,
                                                                                 predict_DT, all_column_dataframe)
                        classification_report_dt = import_text_processing_class.classification_report_dt(y_val,
                                                                                                         predict_val_DT)
                        print("Decision Tree ", classification_report_dt)
                        #logging.info("Decision Tree ", classification_report_dt)
                        predict_val_SVM = import_text_processing_class.validation_svm_model_fit(X_train, y_train,
                                                                                                X_val,
                                                                                                X_train_tfidf)
                        predict_SVM = import_text_processing_class.svm_model_fit(column_name, column_name_modified,
                                                                                 train_column_name,
                                                                                 test_column_name,
                                                                                 X_train_tfidf)
                        score_SVM = import_text_processing_class.svm_model_predict(column_name,
                                                                                   column_name_modified,
                                                                                   train_column_name,
                                                                                   test_column_name,
                                                                                   predict_SVM,
                                                                                   all_column_dataframe)
                        classification_report_svm = import_text_processing_class.classification_report_svm(y_val,
                                                                                                           predict_val_SVM)

                        print("SVM ", classification_report_svm)
                        #logging.info("SVM ", classification_report_svm)
                        predict_val_SVM_GS = import_text_processing_class.validation_svm_grid_search_model_fit(
                            X_train,
                            y_train,
                            X_val,
                            X_train_tfidf)
                        predict_SVM_GS = import_text_processing_class.svm_grid_search_model_fit(column_name,
                                                                                                column_name_modified,
                                                                                                train_column_name,
                                                                                                test_column_name,
                                                                                                X_train_tfidf)
                        score_SVM_GS = import_text_processing_class.svm_grid_search_model_predict(column_name,
                                                                                                  column_name_modified,
                                                                                                  train_column_name,
                                                                                                  test_column_name,
                                                                                                  predict_SVM_GS,
                                                                                                  all_column_dataframe)
                        classification_report_gs_svm = import_text_processing_class.classification_report_gs_svm(
                            y_val,
                            predict_val_SVM_GS)
                        print("SVM GS ", classification_report_gs_svm)
                        #logging.info("SVM GS ", classification_report_gs_svm)
                        predict_val_NB_NLTK = import_text_processing_class.validation_nltk_nb_model_fit(X_train,
                                                                                                        y_train,
                                                                                                        X_val,
                                                                                                        X_train_tfidf)
                        predict_NB_NLTK = import_text_processing_class.nltk_nb_model_fit(column_name,
                                                                                         column_name_modified,
                                                                                         train_column_name,
                                                                                         test_column_name,
                                                                                         X_train_tfidf)
                        score_NB_NLTK = import_text_processing_class.nltk_nb_model_predict(column_name,
                                                                                           column_name_modified,
                                                                                           train_column_name,
                                                                                           test_column_name,
                                                                                           predict_NB_NLTK,
                                                                                           all_column_dataframe)
                        classification_report_nltk_nb = import_text_processing_class.classification_report_nltk_nb(
                            y_val,
                            predict_val_NB_NLTK)
                        print("Naive Bayes with NLTK ", classification_report_nltk_nb)
                        #logging.info("Naive Bayes with NLTK ", classification_report_nltk_nb)
                        #logging.info('Training and Predicting completed')
                        """selecting the model with highest accuracy score"""
                        highest = max(classification_report_nb, classification_report_nltk_nb,
                                      classification_report_gs_svm, classification_report_svm,
                                      classification_report_rf, classification_report_dt)
                        #logging.info('Selected the best model')
                        import_classification_report_class = Classification_Reports()
                        all_column_dataframe.to_excel(
                            os.getcwd() + '\\Data Sets\\' + column_name + '_Prediction\\Prediction_of_all_models.xlsx',
                            index=False)
                        #logging.info('Creating Classification Reports')
                        """Creating Classification reports and confusion matrix for each model outputs"""
                        report_nb = classification_report(y_val, predict_val_NB)
                        matrix_nb = confusion_matrix(y_val, predict_val_NB,
                                                     labels=train_column_name[column_name_modified].unique())
                        print(matrix_nb)
                        #logging.info(matrix_nb)
                        saving_CL_report_nb = import_classification_report_class.classifcation_report_processing(
                            column_name, 'NB', report_nb)
                        print(saving_CL_report_nb)
                        #logging.info(saving_CL_report_nb)
                        report_dt = classification_report(y_val, predict_val_DT)
                        matrix_dt = confusion_matrix(y_val, predict_val_DT,
                                                     labels=train_column_name[column_name_modified].unique())
                        print(matrix_dt)
                        #logging.info(matrix_dt)
                        saving_CL_report_dt = import_classification_report_class.classifcation_report_processing(
                            column_name, 'DT', report_dt)
                        print(saving_CL_report_dt)
                        #logging.info(saving_CL_report_dt)
                        report_rf = classification_report(y_val, predict_val_RF)
                        matrix_rf = confusion_matrix(y_val, predict_val_RF,
                                                     labels=train_column_name[column_name_modified].unique())
                        print(matrix_rf)
                        saving_CL_report_rf = import_classification_report_class.classifcation_report_processing(
                            column_name, 'RF', report_rf)
                        print(saving_CL_report_rf)
                        #logging.info(saving_CL_report_rf)
                        report_svm = classification_report(y_val, predict_val_SVM)
                        matrix_svm = confusion_matrix(y_val, predict_val_SVM,
                                                      labels=train_column_name[column_name_modified].unique())
                        print(matrix_svm)
                        #logging.info(matrix_svm)
                        saving_CL_report_svm = import_classification_report_class.classifcation_report_processing(
                            column_name, 'SVM', report_svm)
                        print(saving_CL_report_svm)
                        #logging.info(saving_CL_report_svm)
                        report_svm_gs = classification_report(y_val, predict_val_SVM_GS)
                        saving_CL_report_svm_gs = import_classification_report_class.classifcation_report_processing(
                            column_name, 'SVM_GS', report_svm_gs)
                        print(saving_CL_report_svm_gs)
                        #logging.info(saving_CL_report_svm_gs)
                        matrix_svm_gs = confusion_matrix(y_val, predict_val_SVM_GS,
                                                         labels=train_column_name[column_name_modified].unique())
                        print(matrix_svm_gs)
                        #logging.info(matrix_svm_gs)
                        report_nltk_nb = classification_report(y_val, predict_val_NB_NLTK)
                        #logging.info(matrix_svm_gs)
                        matrix_nltk_nb = confusion_matrix(y_val, predict_val_NB_NLTK,
                                                          labels=train_column_name[column_name_modified].unique())
                        print(matrix_nltk_nb)
                        #logging.info(matrix_nltk_nb)
                        saving_CL_report_nb_nltk = import_classification_report_class.classifcation_report_processing(
                            column_name, 'NLTK_NB', report_nltk_nb)
                        print(saving_CL_report_nb_nltk)
                        #logging.info(saving_CL_report_nb_nltk)
                        #logging.info('Classification Report written')
                        print("Model with the highest score is selected automatically.")
                        #logging.info("Model with the highest score is selected automatically.")
                        # Rating the prediction with classigication report and confusion matrix from the best model
                        #logging.info('Writing the Prediction file')
                        """scoring with string similarity for the highest performing model"""
                        if classification_report_nb == highest:
                            column_model_name = import_text_processing_class.prediction_rate(column_name,
                                                                                             column_name_modified,
                                                                                             train_column_name,
                                                                                             test_column_name,
                                                                                             'NB',
                                                                                             'Prediction_of_all_models.xlsx')


                        elif classification_report_dt == highest:
                            column_model_name = import_text_processing_class.prediction_rate(column_name,
                                                                                             column_name_modified,
                                                                                             train_column_name,
                                                                                             test_column_name,
                                                                                             'DT',
                                                                                             'Prediction_of_all_models.xlsx')

                        elif classification_report_rf == highest:
                            column_model_name = import_text_processing_class.prediction_rate(column_name,
                                                                                             column_name_modified,
                                                                                             train_column_name,
                                                                                             test_column_name,
                                                                                             'RF',
                                                                                             'Prediction_of_all_models.xlsx')

                        elif classification_report_svm == highest:
                            column_model_name = import_text_processing_class.prediction_rate(column_name,
                                                                                             column_name_modified,
                                                                                             train_column_name,
                                                                                             test_column_name,
                                                                                             'SVM',
                                                                                             'Prediction_of_all_models.xlsx')

                        elif classification_report_gs_svm == highest:
                            column_model_name = import_text_processing_class.prediction_rate(column_name,
                                                                                             column_name_modified,
                                                                                             train_column_name,
                                                                                             test_column_name,
                                                                                             'SVM_GS',
                                                                                             'Prediction_of_all_models.xlsx')

                        elif classification_report_nltk_nb == highest:
                            column_model_name = import_text_processing_class.prediction_rate(column_name,
                                                                                             column_name_modified,
                                                                                             train_column_name,
                                                                                             test_column_name,
                                                                                             'NLTK_NB',
                                                                                             'Prediction_of_all_models.xlsx')

                        column_prediction_sorting_order = '0'
                        column_prediction_sorting_message = '0'
                        #logging.info('Prediction file written')
                        while (column_prediction_sorting_order == '0'):
                            """sorting the prediction file before saving based on the user input from best to 
                            worst or vice-versa"""
                            column_prediction_sorting_order = column_prediction_sorting_order_input
                            if (column_prediction_sorting_order == '1'):
                                column_model_name = column_model_name.sort_values('F1_Score')
                            elif (column_prediction_sorting_order == '2'):
                                column_model_name = column_model_name.sort_values('F1_Score', ascending=False)
                            else:
                                column_prediction_sorting_message_colored = colored(
                                    "The choice you made does not exist. Do you wish to try again.\nPress 0 to try again or any key to exit\nYour Response: ",
                                    'blue')
                                column_prediction_sorting_message = '0'
                                if (column_prediction_sorting_message == '0'):
                                    column_prediction_sorting_order = '0'
                                else:
                                    column_prediction_sorting_order = '3'
                                    print('Choice not found, file will be saved without sorting.')
                        file_to_save = pd.DataFrame()
                        column_model_name = column_model_name.drop(['Prediction_Percent'], axis=1)
                        file_to_save = column_model_name.copy()
                        #logging.info('Saving the progress')
                        """Creating new directory to save all the predictions combined"""
                        input_file_name_folder_for_columns = input_file_name.rsplit('\\', 1)[1]
                        path_to_save = os.getcwd() + '\\Data Sets\\' + column_name + '_Prediction\\Predicted_Score_' + column_name + '.xlsx'
                        if not os.path.exists(os.getcwd() + '/Data Sets/Combined_all_column_output'):
                            os.makedirs(os.getcwd() + '/Data Sets/Combined_all_column_output')
                        if not os.path.exists(
                                os.getcwd() + '/Data Sets/Combined_all_column_output/' + input_file_name_folder_for_columns):
                            os.makedirs(
                                os.getcwd() + '/Data Sets/Combined_all_column_output/' + input_file_name_folder_for_columns)
                        path_to_save_combined_all_column_output = os.getcwd() + '\\Data Sets\\Combined_all_column_output\\' + input_file_name_folder_for_columns + '\\Predicted_Score_all_columns.xlsx'


                        save_combined_all_column_output = column_model_name.copy()
                        cleanedframefilepath = os.getcwd() + '\\Data Sets\\Cleaned_File.xlsx'
                        jsonframefilepath = os.getcwd() + '\\Data Sets\\Json_Mapping file.xlsx'

                        saver = Save_With_All_Values()
                        saver.saving_all_values(save_combined_all_column_output, cleanedframefilepath,
                                                jsonframefilepath, input_file_name_folder_for_columns)

                        """Saving all the predicted models together in one file"""
                        if not os.path.exists(path_to_save_combined_all_column_output):
                            with pd.ExcelWriter(path_to_save_combined_all_column_output,
                                                engine='xlsxwriter') as writer:
                                save_combined_all_column_output.to_excel(writer, column_name, index=False)
                                writer.save()
                        else:
                            book = load_workbook(path_to_save_combined_all_column_output)
                            with pd.ExcelWriter(path_to_save_combined_all_column_output,
                                                engine='openpyxl') as writer:
                                writer.book = book
                                writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
                                save_combined_all_column_output.to_excel(writer, column_name, index=False)
                                writer.save()

                        file_to_save.to_excel(path_to_save, index=False)
                        file_to_save.head()

                        path_to_save_combined_all_values = os.getcwd() + '\\Data ' \
                                                                        'Sets\\Combined_all_column_output\\' + input_file_name_folder_for_columns + '\\Prediction_with_Orginal_Values.xlsx'

                        # saverjson = Save_as_JSON()
                        # saverjson.saving_json(path_to_save_combined_all_values, jsonframefilepath, input_file_name_folder_for_columns)
                    #else:
                        #logging.info("The column name entered doesnot exists in Input file, please check again..")
                #logging.info("The process finished for the given column..")
                    print(colored("The process finished for the give column.", 'green'))
                else:
                    #logging.info('Input file not found')
                    print(colored("The file at the give path does not exist.  Line Number 1779", 'red'))
                    input_file_path_message_colored = colored(
                        "Do you wish to try again.\nPress 0 to try again or any key to exit\nYour Response: ", 'blue')
                    input_file_path_message = 1
                    if (input_file_path_message == '0'):
                        input_file_path = '0'
                    else:
                        break
                response_message = colored(
                    "Press Any Key To finish the process.\nPress 1: To do Machine Learning on another column.\nYour Response: ",
                    'blue')
                response = 2
                if (response == '1'):
                    input_file_path = '0'
                else:
                    pass

# main method to run the Text processing

def main():
    """main method that calls the method to create ui and do the rest of the processing"""
    input_file_name = sys.argv[1]
    input_file_sheet_name = sys.argv[2]

    print('input_file_name: ', input_file_name)
    print('input_file_sheet_name: ', input_file_sheet_name)
    object_process_all_text_columns = Process_All_Text_Columns()
    object_process_all_text_columns.ml_process_general_text_columns_processing(input_file_name,
                                                                               input_file_sheet_name)


if __name__ == '__main__':
    main()
